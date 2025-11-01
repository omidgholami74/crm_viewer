import sys
import sqlite3
import pandas as pd
import re
import logging
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTableWidget, QTableWidgetItem, QHeaderView, QProgressBar, QMessageBox,
    QFileDialog, QLabel, QDialog, QComboBox, QPushButton, QListWidget, QListWidgetItem, QLineEdit,
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal,QTimer
from PyQt5.QtGui import QFont, QPixmap, QColor
from pyqtgraph import PlotWidget, mkPen
from qfluentwidgets import (
    ComboBox, LineEdit, PrimaryPushButton, CheckBox, CardWidget,
    setTheme, Theme, FluentIcon, TitleLabel
)
from persiantools.jdatetime import JalaliDate
from PyQt5.QtGui import QFont, QColor
from PyQt5.QtCore import Qt
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import numpy as np
from pathlib import Path
from PIL import Image
import csv
import shutil
import os

# Setup logging with UTF-8 encoding
log_file = Path("crm_visualizer.log").resolve()
file_handler = logging.FileHandler(log_file, mode='w', encoding='utf-8')
file_handler.setLevel(logging.DEBUG)
file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))

console_handler = logging.StreamHandler(sys.stdout)
console_handler.setLevel(logging.DEBUG)
console_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))

logger = logging.getLogger()
logger.setLevel(logging.DEBUG)
logger.handlers = []
logger.addHandler(file_handler)
logger.addHandler(console_handler)

def normalize_crm_id(crm_id):
    """Extract numeric part from CRM ID (e.g., 'CRM 258b' → '258', '258 b' → '258')."""
    if not isinstance(crm_id, str):
        return None
    crm_pattern = re.compile(r'^(?:\s*CRM\s*)?(\d{3})(?:\s*[a-zA-Z])?$', re.IGNORECASE)
    match = crm_pattern.match(crm_id.strip())
    if match:
        # logger.debug(f"Normalized CRM ID: {crm_id} → {match.group(1)}")
        return match.group(1)
    # logger.debug(f"Invalid CRM ID format: {crm_id}")
    return None

def validate_jalali_date(date_str):
    """Validate Jalali date string (YYYY/MM/DD)."""
    try:
        if not isinstance(date_str, str):
            return False
        # اطمینان از فرمت YYYY/MM/DD
        parts = date_str.split('/')
        if len(parts) != 3:
            return False
        year, month, day = map(int, parts)
        # تبدیل روز و ماه به فرمت دو رقمی برای مقایسه
        date_str_normalized = f"{year:04d}/{month:02d}/{day:02d}"
        JalaliDate(year, month, day)
        return date_str_normalized
    except (ValueError, TypeError):
        return False

def validate_percentage(text):
    """Validate percentage input (must be positive float)."""
    try:
        value = float(text)
        return value > 0
    except (ValueError, TypeError):
        return False

def split_element_name(element):
    """Split element name like 'Ce140' into 'Ce 140'."""
    if not isinstance(element, str):
        return element
    match = re.match(r'^([A-Za-z]+)(\d+\.?\d*)$', element.strip())
    if match:
        symbol, number = match.groups()
        return f"{symbol} {number}"
    return element

def extract_date(file_name):
    """Extract date from file_name like '1404-07-01' or '1404-07-1'."""
    try:
        match = re.match(r'(\d{4}-\d{2}-\d{1,2})', file_name)
        if match:
            date_str = match.group(1)
            year, month, day = map(int, date_str.split('-'))
            date_str = f"{year:04d}/{month:02d}/{day:02d}"
            JalaliDate(year, month, day)  # Validate Jalali date
            return date_str
        logger.warning(f"No valid date found in filename: {file_name}")
        return None
    except Exception as e:
        logger.error(f"Error extracting date from {file_name}: {str(e)}")
        return None

def is_numeric(value):
    """Check if a value can be converted to float."""
    if value is None or str(value).strip() == "":
        return False
    try:
        float(value)
        return True
    except ValueError:
        return False

def vacuum_db(db_path):
    """Run VACUUM on the database to reclaim space."""
    try:
        conn = sqlite3.connect(db_path)
        conn.execute("VACUUM")
        conn.commit()
        conn.close()
        logger.info(f"Database vacuumed successfully: {db_path}")
    except Exception as e:
        logger.error(f"Error vacuuming database {db_path}: {str(e)}")

def load_raw_file(file_path, db_path, selected_device=None):
    """Load and parse raw CSV/.rep file into a DataFrame with required columns."""
    file_path = Path(file_path)
    logger.info(f"Processing raw file: {file_path} with device: {selected_device}")
    allowed_crms = ['258', '252', '906', '506', '233', '255', '263', '260']
    crm_pattern = re.compile(r'^(?:\s*CRM\s*)?(\d{3}(?:\s*[a-zA-Z])?)$', re.IGNORECASE)
    blank_pattern = re.compile(r'(?:CRM\s*)?(?:BLANK|BLNK)(?:S|s)?(?:\s+.*)?', re.IGNORECASE)

    # Check for duplicate file
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM crm_data WHERE file_name = ?", (file_path.name,))
    if cursor.fetchone()[0] > 0:
        conn.close()
        logger.warning(f"File {file_path.name} already exists in database")
        raise ValueError(f"File {file_path.name} already exists in the database")
    conn.close()

    try:
        is_new_format = False
        with open(file_path, 'r', encoding='utf-8') as f:
            preview_lines = [f.readline().strip() for _ in range(10)]
            logger.debug(f"CSV preview (first 10 lines) for {file_path.name}:\n{preview_lines}")
            is_new_format = any("Sample ID:" in line for line in preview_lines) or \
                            any("Net Intensity" in line for line in preview_lines)
        logger.info(f"File {file_path.name} detected as {'new' if is_new_format else 'old'} format")

        data_rows = []
        file_date = extract_date(file_path.name)
        if is_new_format:
            current_sample = None
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = list(csv.reader(f, delimiter=',', quotechar='"'))
                total_rows = len(reader)
                logger.debug(f"Total rows in CSV {file_path.name}: {total_rows}")
                for idx, row in enumerate(reader):
                    if idx == total_rows - 1:
                        logger.debug(f"Skipping last row {idx} in {file_path.name}")
                        continue
                    if not row or all(cell.strip() == "" for cell in row):
                        continue
                    
                    if len(row) > 0 and row[0].startswith("Sample ID:"):
                        current_sample = row[1].strip() if len(row) > 1 else "Unknown_Sample"
                        continue
                    
                    if len(row) > 0 and (row[0].startswith("Method File:") or row[0].startswith("Calibration File:")):
                        continue
                    
                    if current_sample is None:
                        current_sample = "Unknown_Sample"
                        logger.warning(f"No Sample ID found before row {idx}, using default: {current_sample}")
                    
                    element = split_element_name(row[0].strip())
                    try:
                        concentration = float(row[5]) if len(row) > 5 and is_numeric(row[5]) else None
                        if concentration is not None:
                            blank_match = blank_pattern.match(current_sample)
                            crm_match = crm_pattern.match(current_sample)
                            if blank_match:
                                crm_id_value = "BLANK"  # برای BLANK، crm_id هم BLANK است
                                norm_crm_id = None
                            elif crm_match:
                                # استخراج فقط عدد از current_sample
                                crm_id_value = normalize_crm_id(current_sample)
                                norm_crm_id = crm_id_value
                                if norm_crm_id not in allowed_crms:
                                    logger.debug(f"Skipping row {idx} in {file_path.name}: Invalid CRM ID {norm_crm_id}")
                                    continue
                            else:
                                logger.debug(f"Skipping row {idx} in {file_path.name}: Invalid type {current_sample}")
                                continue
                            
                            # current_sample به عنوان solution_label (کل عبارت) و crm_id_value به عنوان crm_id (فقط عدد)
                            data_rows.append({
                                "crm_id": crm_id_value,  # فقط عدد یا BLANK
                                "solution_label": current_sample,  # کل عبارت اصلی
                                "element": element,
                                "value": concentration,
                                "file_name": file_path.name,
                                "folder_name": selected_device or str(file_path.parent.name),
                                "date": file_date
                            })
                            logger.debug(f"Added row {idx} in {file_path.name}: crm_id={crm_id_value}, solution_label={current_sample}, norm_crm_id={norm_crm_id}")
                        else:
                            logger.warning(f"Skipping row {idx} in {file_path.name}: Non-numeric concentration")
                    except Exception as e:
                        logger.error(f"Error processing row {idx} in {file_path.name}: {str(e)}")
                        continue
        else:
            # Old format
            temp_df = pd.read_csv(file_path, header=None, nrows=1, encoding='utf-8', low_memory=False)
            logger.debug(f"CSV header preview for {file_path.name}: {temp_df.to_string()}")
            if temp_df.iloc[0].notna().sum() == 1:
                df = pd.read_csv(file_path, header=1, encoding='utf-8', on_bad_lines='skip', low_memory=False)
            else:
                df = pd.read_csv(file_path, header=0, encoding='utf-8', on_bad_lines='skip', low_memory=False)
            logger.debug(f"Loaded CSV {file_path.name} with {len(df)} rows")
            
            df = df.iloc[:-1]
            logger.debug(f"Removed last row, remaining rows: {len(df)}")
            
            expected_columns = ["Solution Label", "Element", "Corr Con"]
            column_mapping = {"Sample ID": "Solution Label"}
            df.rename(columns=column_mapping, inplace=True)
            
            if not all(col in df.columns for col in expected_columns):
                missing_cols = set(expected_columns) - set(df.columns)
                logger.error(f"Required columns missing in {file_path.name}: {', '.join(missing_cols)}")
                raise ValueError(f"Required columns missing: {', '.join(missing_cols)}")
            
            df['Element'] = df['Element'].apply(split_element_name)
            
            # اصلاح بخش مهم: استخراج crm_id درست
            df['norm_crm_id'] = df['Solution Label'].apply(normalize_crm_id)
            df['is_blank'] = df['Solution Label'].apply(lambda x: bool(blank_pattern.match(str(x).strip())))
            
            # برای BLANK ها، crm_id = "BLANK"
            # برای CRM ها، crm_id = norm_crm_id (فقط عدد)
            # solution_label = Solution Label اصلی (کل عبارت)
            df['crm_id'] = df.apply(
                lambda row: "BLANK" if row['is_blank'] else (row['norm_crm_id'] if pd.notna(row['norm_crm_id']) else None),
                axis=1
            )
            
            # فیلتر کردن رکوردها
            df = df[(df['is_blank']) | (df['norm_crm_id'].isin(allowed_crms))]
            
            df['value'] = pd.to_numeric(df['Corr Con'], errors='coerce')
            df = df.dropna(subset=['value'])
            df['file_name'] = file_path.name
            df['folder_name'] = selected_device or str(file_path.parent.name)
            df['date'] = file_date
            
            # انتخاب ستون‌های نهایی
            data_rows = df[['crm_id', 'Solution Label', 'Element', 'value', 'file_name', 'folder_name', 'date']].rename(
                columns={'Solution Label': 'solution_label', 'Element': 'element'}
            ).to_dict('records')
            
            logger.debug(f"Old format processing - Unique crm_id: {set(row['crm_id'] for row in data_rows if row['crm_id'])}")
            logger.debug(f"Old format processing - Sample data_rows: {data_rows[:2]}")
        
        if not data_rows:
            logger.error(f"No valid data found in {file_path.name}")
            raise ValueError("No valid data found in the file")
        
        df = pd.DataFrame(data_rows)
        df = df[['crm_id', 'solution_label', 'element', 'value', 'file_name', 'folder_name', 'date']]
        logger.debug(f"Final DataFrame columns: {df.columns.tolist()}")
        logger.debug(f"Final DataFrame sample:\n{df.head().to_string()}")
        logger.debug(f"Unique crm_id values: {df['crm_id'].unique()}")
        logger.info(f"Successfully processed {file_path.name} with {len(df)} rows")
        return df
    
    except Exception as e:
        logger.error(f"Error loading {file_path}: {str(e)}")
        raise

class DeviceSelectionDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Select Device")
        self.setFixedSize(300, 150)
        
        self.layout = QVBoxLayout()
        self.label = QLabel("Please select the device type for the imported file:")
        self.device_combo = QComboBox()
        self.device_combo.addItems(['mass', 'oes 4ac', 'oes fire'])
        self.button_layout = QHBoxLayout()
        self.ok_button = QPushButton("OK")
        self.cancel_button = QPushButton("Cancel")
        
        self.button_layout.addWidget(self.ok_button)
        self.button_layout.addWidget(self.cancel_button)
        self.layout.addWidget(self.label)
        self.layout.addWidget(self.device_combo)
        self.layout.addLayout(self.button_layout)
        self.setLayout(self.layout)
        
        self.ok_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)

    def get_device(self):
        return self.device_combo.currentText()

class LoadDeleteFilesDialogThread(QThread):
    dialog_ready = pyqtSignal(list, dict)  # list of file_names, dict of record_counts
    error_occurred = pyqtSignal(str)
    progress_updated = pyqtSignal(int)

    def __init__(self, db_path, file_names):
        super().__init__()
        self.db_path = db_path
        self.file_names = list(set(file_names))  # Remove duplicates upfront
        self.record_counts = {}

    def run(self):
        try:
            logger.info(f"Loading {len(self.file_names)} unique files for deletion dialog")
            self.progress_updated.emit(10)
            
            # Single database connection for all queries
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            total_files = len(self.file_names)
            processed = 0
            
            # Use IN clause for better performance instead of individual queries
            placeholders = ','.join(['?'] * total_files)
            cursor.execute(f"""
                SELECT file_name, COUNT(*) as record_count 
                FROM crm_data 
                WHERE file_name IN ({placeholders})
                GROUP BY file_name
            """, self.file_names)
            
            results = cursor.fetchall()
            self.progress_updated.emit(80)
            
            # Create record_counts dictionary from results
            for file_name, count in results:
                self.record_counts[file_name] = count
            
            # For files not found in results, count is 0
            for file_name in self.file_names:
                if file_name not in self.record_counts:
                    self.record_counts[file_name] = 0
            
            conn.close()
            self.progress_updated.emit(100)
            
            logger.info(f"Loaded record counts for {len(self.record_counts)} files")
            self.dialog_ready.emit(sorted(self.file_names), self.record_counts)
            
        except Exception as e:
            logger.error(f"Error loading delete files dialog: {str(e)}")
            self.error_occurred.emit(f"Failed to load files for deletion: {str(e)}")
            self.progress_updated.emit(100)

class DeleteFilesDialog(QDialog):
    def __init__(self, parent=None, file_names=None, db_path=None):
        super().__init__(parent)
        self.setWindowTitle("Delete Files")
        self.setFixedSize(500, 500)
        self.db_path = db_path
        self.record_counts = {}
        
        self.layout = QVBoxLayout()
        
        # Progress bar and status label
        self.progress_bar = QProgressBar()
        self.progress_bar.setMaximum(100)
        self.progress_bar.setVisible(False)
        
        self.status_label = QLabel("Loading files... Please wait.")
        self.status_label.setAlignment(Qt.AlignCenter)
        
        # Main content
        self.main_label = QLabel("Select files to delete (number of records shown):")
        self.file_list = QListWidget()
        
        self.button_layout = QHBoxLayout()
        self.delete_button = QPushButton("Delete Selected")
        self.cancel_button = QPushButton("Cancel")
        self.refresh_button = QPushButton("Refresh")
        
        self.button_layout.addWidget(self.delete_button)
        self.button_layout.addWidget(self.refresh_button)
        self.button_layout.addWidget(self.cancel_button)
        
        # Add widgets to layout
        self.layout.addWidget(self.status_label)
        self.layout.addWidget(self.progress_bar)
        self.layout.addWidget(self.main_label)
        self.layout.addWidget(self.file_list)
        self.layout.addLayout(self.button_layout)
        self.setLayout(self.layout)
        
        # Connections
        self.delete_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)
        self.refresh_button.clicked.connect(self.refresh_files)
        
        # Load files if provided and non-empty, otherwise show empty message
        if file_names is not None and len(file_names) > 0:  # Fix: Check length explicitly
            self.load_files_async(file_names)
        else:
            self.status_label.setText("No files provided")
            self.progress_bar.setVisible(False)
            self.delete_button.setEnabled(False)
            self.refresh_button.setEnabled(False)
    
    def load_files_async(self, file_names):
        """Load files asynchronously with progress"""
        self.progress_bar.setVisible(True)
        self.delete_button.setEnabled(False)
        self.cancel_button.setEnabled(False)
        self.refresh_button.setEnabled(False)
        self.file_list.clear()
        
        self.loader_thread = LoadDeleteFilesDialogThread(self.db_path, file_names)
        self.loader_thread.dialog_ready.connect(self.on_files_loaded)
        self.loader_thread.error_occurred.connect(self.on_load_error)
        self.loader_thread.progress_updated.connect(self.progress_bar.setValue)
        self.loader_thread.finished.connect(self.on_loading_finished)
        self.loader_thread.start()
    
    def on_files_loaded(self, file_names, record_counts):
        """Handle successful file loading"""
        self.file_names = file_names
        self.record_counts = record_counts
        
        # Populate file list
        self.file_list.clear()
        for file_name in file_names:
            count = record_counts.get(file_name, 0)
            item = QListWidgetItem(f"{file_name} ({count:,} records)")
            item.setData(Qt.UserRole, file_name)
            item.setCheckState(Qt.Unchecked)
            item.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
            self.file_list.addItem(item)
        
        logger.info(f"Populated dialog with {len(file_names)} files")
    
    def on_load_error(self, error_message):
        """Handle loading error"""
        self.status_label.setText(f"Error: {error_message}")
        logger.error(f"Delete dialog load error: {error_message}")
        QMessageBox.critical(self, "Error", error_message)
    
    def on_loading_finished(self):
        """Enable UI after loading completes"""
        self.progress_bar.setVisible(False)
        self.delete_button.setEnabled(True)
        self.cancel_button.setEnabled(True)
        self.refresh_button.setEnabled(True)
        self.status_label.setText(f"Loaded {len(self.record_counts)} files")
    
    def refresh_files(self):
        """Refresh file list"""
        if hasattr(self, 'file_names') and self.file_names:
            self.load_files_async(self.file_names)
    
    def get_selected_files(self):
        """Get selected files for deletion"""
        selected = []
        for i in range(self.file_list.count()):
            item = self.file_list.item(i)
            if item.checkState() == Qt.Checked:
                selected.append(item.data(Qt.UserRole))
        return selected
    
    def get_total_records(self):
        """Get total records for selected files"""
        total = 0
        for file_name in self.get_selected_files():
            total += self.record_counts.get(file_name, 0)
        return total
    
def delete_files(self):
    """Updated delete_files method with async loading"""
    all_df = pd.concat([self.crm_df, self.blank_df])
    file_names = all_df['file_name'].unique().tolist()  # Convert to list explicitly
    
    if not file_names:  # Fix: Check if list is empty
        QMessageBox.warning(self, "Warning", "No files to delete")
        return
    
    # Show dialog with async loading
    dialog = DeleteFilesDialog(self, file_names, self.crm_db_path)
    
    # Connect dialog signals to handle completion
    if dialog.exec_() == QDialog.Accepted:
        selected_files = dialog.get_selected_files()
        if not selected_files:
            QMessageBox.warning(self, "Warning", "No files selected for deletion")
            return
        
        total_records = dialog.get_total_records()
        confirm = QMessageBox.question(
            self, "Confirm Delete",
            f"Are you sure you want to delete {len(selected_files)} files with {total_records:,} records?\n\n"
            f"Selected files:\n{chr(10).join(selected_files[:5])}{'...' if len(selected_files) > 5 else ''}",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if confirm == QMessageBox.Yes:
            self.progress_bar.setVisible(True)
            self.delete_thread = DeleteFilesThread(self.crm_db_path, selected_files)
            self.delete_thread.delete_completed.connect(self.on_delete_completed)
            self.delete_thread.error_occurred.connect(self.on_data_error)
            self.delete_thread.progress_updated.connect(self.progress_bar.setValue)
            self.delete_thread.finished.connect(lambda: self.progress_bar.setVisible(False))
            self.delete_thread.start()
class EditRecordDialog(QDialog):
    def __init__(self, parent=None, record=None, db_path=None):
        super().__init__(parent)
        self.setWindowTitle("Edit Record")
        self.setFixedSize(400, 300)
        self.record = record
        self.db_path = db_path
        
        self.layout = QVBoxLayout()
        
        self.crm_id_label = QLabel("CRM ID:")
        self.crm_id_edit = LineEdit()
        self.crm_id_edit.setText(str(record['crm_id']) if pd.notna(record['crm_id']) else "")
        
        self.solution_label = QLabel("Solution Label:")
        self.solution_edit = LineEdit()
        self.solution_edit.setText(str(record['solution_label']) if pd.notna(record['solution_label']) else "")
        
        self.element_label = QLabel("Element:")
        self.element_edit = LineEdit()
        self.element_edit.setText(str(record['element']) if pd.notna(record['element']) else "")
        
        self.value_label = QLabel("Value:")
        self.value_edit = LineEdit()
        self.value_edit.setText(f"{record['value']:.2f}" if pd.notna(record['value']) else "")
        
        self.date_label = QLabel("Date (YYYY/MM/DD):")
        self.date_edit = LineEdit()
        self.date_edit.setText(str(record['date']) if pd.notna(record['date']) else "")
        
        self.button_layout = QHBoxLayout()
        self.save_button = QPushButton("Save")
        self.cancel_button = QPushButton("Cancel")
        
        self.button_layout.addWidget(self.save_button)
        self.button_layout.addWidget(self.cancel_button)
        self.layout.addWidget(self.crm_id_label)
        self.layout.addWidget(self.crm_id_edit)
        self.layout.addWidget(self.solution_label)
        self.layout.addWidget(self.solution_edit)
        self.layout.addWidget(self.element_label)
        self.layout.addWidget(self.element_edit)
        self.layout.addWidget(self.value_label)
        self.layout.addWidget(self.value_edit)
        self.layout.addWidget(self.date_label)
        self.layout.addWidget(self.date_edit)
        self.layout.addLayout(self.button_layout)
        self.setLayout(self.layout)
        
        self.save_button.clicked.connect(self.accept)
        self.cancel_button.clicked.connect(self.reject)

    def get_updated_record(self):
        return {
            'crm_id': self.crm_id_edit.text(),
            'solution_label': self.solution_edit.text(),
            'element': self.element_edit.text(),
            'value': float(self.value_edit.text()) if is_numeric(self.value_edit.text()) else self.record['value'],
            'date': self.date_edit.text() if validate_jalali_date(self.date_edit.text()) else self.record['date']
        }

class DataLoaderThread(QThread):
    data_loaded = pyqtSignal(pd.DataFrame, pd.DataFrame)
    error_occurred = pyqtSignal(str)
    progress_updated = pyqtSignal(int)

    def __init__(self, db_path):
        super().__init__()
        self.db_path = db_path

    def run(self):
        try:
            logger.debug(f"Loading data from {self.db_path}")
            self.progress_updated.emit(20)
            conn = sqlite3.connect(self.db_path)
            df = pd.read_sql_query("SELECT * FROM crm_data", conn)
            conn.close()
            self.progress_updated.emit(60)

            if 'date' not in df.columns or df['date'].isna().all():
                df['date'] = df['file_name'].apply(extract_date)
            df = df.dropna(subset=['date'])
            df['year'] = df['date'].apply(lambda x: int(x.split('/')[0]) if pd.notna(x) else 0)
            df['month'] = df['date'].apply(lambda x: int(x.split('/')[1]) if pd.notna(x) else 0)
            df['day'] = df['date'].apply(lambda x: int(x.split('/')[2]) if pd.notna(x) else 0)
            self.progress_updated.emit(80)

            crm_df = df[df['crm_id'] != 'BLANK'].copy()
            blank_df = df[df['crm_id'] == 'BLANK'].copy()
            crm_df['norm_crm_id'] = crm_df['crm_id'].apply(normalize_crm_id)
            allowed_crms = ['258', '252', '906', '506', '233', '255', '263', '260']
            crm_df = crm_df[crm_df['norm_crm_id'].isin(allowed_crms)].dropna(subset=['norm_crm_id'])
            logger.debug(f"Sample CRM dates: {crm_df['date'].head(10).to_list()}")
            logger.debug(f"Sample BLANK dates: {blank_df['date'].head(10).to_list()}")
            logger.debug(f"Unique norm_crm_id values: {crm_df['norm_crm_id'].unique()}")
            self.progress_updated.emit(100)
            logger.info(f"Loaded {len(crm_df)} CRM records and {len(blank_df)} BLANK records from {self.db_path}")
            self.data_loaded.emit(crm_df, blank_df)
        except Exception as e:
            logger.error(f"Data loading error: {str(e)}")
            self.error_occurred.emit(f"Failed to load data: {str(e)}")

class ImportFileThread(QThread):
    import_completed = pyqtSignal(pd.DataFrame)
    error_occurred = pyqtSignal(str)
    progress_updated = pyqtSignal(int)

    def __init__(self, file_path, db_path, selected_device=None):
        super().__init__()
        self.file_path = file_path
        self.db_path = db_path
        self.selected_device = selected_device

    def run(self):
        try:
            self.progress_updated.emit(20)
            file_path = Path(self.file_path)
            ext = file_path.suffix.lower()

            if ext == '.rep':
                csv_path = file_path.with_suffix('.csv')
                if not csv_path.exists():
                    shutil.copy(file_path, csv_path)
                    logger.debug(f"Converted {file_path} to {csv_path}")
                file_path = csv_path
                ext = '.csv'

            if ext != '.csv':
                raise ValueError("Unsupported file format. Only CSV and .rep are allowed.")

            df = load_raw_file(file_path, self.db_path, self.selected_device)
            self.progress_updated.emit(50)

            conn = sqlite3.connect(self.db_path)
            df.to_sql('crm_data', conn, if_exists='append', index=False)
            conn.commit()
            conn.close()
            vacuum_db(self.db_path)
            self.progress_updated.emit(100)
            logger.info(f"Imported {len(df)} records from {file_path} to {self.db_path} with device {self.selected_device}")
            logger.debug(f"Imported data sample:\n{df.head().to_string()}")
            logger.debug(f"Imported CRM IDs: {df['crm_id'].unique().tolist()}")
            self.import_completed.emit(df)
        except Exception as e:
            logger.error(f"Import error: {str(e)}")
            self.error_occurred.emit(f"Failed to import file: {str(e)}")

class DeleteFilesThread(QThread):
    delete_completed = pyqtSignal()
    error_occurred = pyqtSignal(str)
    progress_updated = pyqtSignal(int)

    def __init__(self, db_path, selected_files):
        super().__init__()
        self.db_path = db_path
        self.selected_files = selected_files

    def run(self):
        try:
            self.progress_updated.emit(20)
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            placeholders = ','.join(['?'] * len(self.selected_files))
            cursor.execute(f"DELETE FROM crm_data WHERE file_name IN ({placeholders})", self.selected_files)
            conn.commit()
            deleted_rows = cursor.rowcount
            conn.close()
            vacuum_db(self.db_path)
            self.progress_updated.emit(100)
            logger.info(f"Deleted {deleted_rows} records for files: {self.selected_files}")
            self.delete_completed.emit()
        except Exception as e:
            logger.error(f"Delete error: {str(e)}")
            self.error_occurred.emit(f"Failed to delete files: {str(e)}")

class FilterThread(QThread):
    filtered_data = pyqtSignal(pd.DataFrame, pd.DataFrame)
    progress_updated = pyqtSignal(int)

    def __init__(self, crm_df, blank_df, filters):
        super().__init__()
        self.crm_df = crm_df
        self.blank_df = blank_df
        self.filters = filters

    def run(self):
        filtered_crm_df = self.crm_df.copy()
        filtered_blank_df = self.blank_df.copy()

        # Device
        if self.filters['device']:
            filtered_crm_df = filtered_crm_df[filtered_crm_df['folder_name'].str.contains(self.filters['device'], case=False, na=False)]
            filtered_blank_df = filtered_blank_df[filtered_blank_df['folder_name'].str.contains(self.filters['device'], case=False, na=False)]

        # CRM ID
        if self.filters['crm']:
            filtered_crm_df = filtered_crm_df[filtered_crm_df['norm_crm_id'] == self.filters['crm']]

        # Element
        if self.filters['element']:
            base_element = self.filters['element']
            filtered_crm_df = filtered_crm_df[
                filtered_crm_df['element'].str.startswith(base_element + ' ', na=False) |
                (filtered_crm_df['element'] == base_element)
            ]
            filtered_blank_df = filtered_blank_df[
                filtered_blank_df['element'].str.startswith(base_element + ' ', na=False) |
                (filtered_blank_df['element'] == base_element)
            ]

        # Date
        if self.filters['from_date']:
            filtered_crm_df = filtered_crm_df[filtered_crm_df['date'] >= self.filters['from_date'].strftime("%Y/%m/%d")]
            filtered_blank_df = filtered_blank_df[filtered_blank_df['date'] >= self.filters['from_date'].strftime("%Y/%m/%d")]
        if self.filters['to_date']:
            filtered_crm_df = filtered_crm_df[filtered_crm_df['date'] <= self.filters['to_date'].strftime("%Y/%m/%d")]
            filtered_blank_df = filtered_blank_df[filtered_blank_df['date'] <= self.filters['to_date'].strftime("%Y/%m/%d")]

        self.filtered_data.emit(filtered_crm_df, filtered_blank_df)

class OutOfRangeFilesDialog(QDialog):
    def __init__(self, parent=None, file_names=[], db_path=None, percentage=10.0, ver_db_path=None):
        super().__init__(parent)
        self.setWindowTitle("Out of Range Elements")
        self.setFixedSize(400, 400)
        self.db_path = db_path
        self.ver_db_path = ver_db_path
        self.percentage = percentage
        
        self.layout = QVBoxLayout()
        self.label = QLabel("Select a file to view out-of-range elements:")
        self.file_list = QListWidget()
        
        for file_name in sorted(set(file_names)):
            item = QListWidgetItem(file_name)
            item.setData(Qt.UserRole, file_name)
            self.file_list.addItem(item)
        
        self.layout.addWidget(self.label)
        self.layout.addWidget(self.file_list)
        self.setLayout(self.layout)
        
        self.file_list.itemClicked.connect(self.on_file_clicked)

    def on_file_clicked(self, item):
        file_name = item.data(Qt.UserRole)
        self.progress_bar = QProgressBar(self)
        self.progress_bar.setMaximum(100)
        self.progress_bar.setVisible(True)
        self.layout.addWidget(self.progress_bar)
        
        self.out_of_range_thread = OutOfRangeThread(self.db_path, file_name, self.percentage, self.ver_db_path)
        self.out_of_range_thread.out_of_range_data.connect(self.on_out_of_range_data)
        self.out_of_range_thread.progress_updated.connect(self.progress_bar.setValue)
        self.out_of_range_thread.finished.connect(lambda: self.progress_bar.setVisible(False))
        self.out_of_range_thread.start()

    def on_out_of_range_data(self, out_df):
        dialog = OutOfRangeTableDialog(self, out_df)
        dialog.exec_()

class OutOfRangeTableDialog(QDialog):
    def __init__(self, parent=None, out_df=None):
        super().__init__(parent)
        self.setWindowTitle("Out of Range Elements")
        self.setMinimumSize(800, 500)
        self.setStyleSheet("""
            QDialog {
                background-color: #F5F6FA;
                font-family: 'Segoe UI', Arial, sans-serif;
            }
            QTableWidget {
                background-color: #FFFFFF;
                border: 1px solid #E0E0E0;
                border-radius: 8px;
                gridline-color: #E0E0E0;
            }
            QTableWidget::item {
                padding: 8px;
            }
            QHeaderView::section {
                background-color: #0078D4;
                color: #FFFFFF;
                border: 1px solid #E0E0E0;
                padding: 10px;
                font-weight: bold;
                font-size: 16px;
                font-family: 'Segoe UI';
            }
        """)

        self.out_df = out_df
        self.layout = QVBoxLayout()
        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(6)  # Added column for Percentage Difference
        self.table_widget.setHorizontalHeaderLabels([
            "CRM ID", "Element", "Value", "Corrected Value", "Ref Value", "Diff %"
        ])

        header = self.table_widget.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Interactive)
        header.setMinimumSectionSize(100)
        header.setDefaultSectionSize(150)
        header.setStretchLastSection(True)
        header.setFont(QFont("Segoe UI", 12, QFont.Bold))

        vertical_header = self.table_widget.verticalHeader()
        vertical_header.setSectionResizeMode(QHeaderView.Interactive)
        vertical_header.setDefaultSectionSize(40)
        vertical_header.setFont(QFont("Segoe UI", 10))

        if out_df is not None and not out_df.empty:
            self.table_widget.setRowCount(len(out_df))
            for i, row in out_df.iterrows():
                crm_id_item = QTableWidgetItem(str(row['crm_id']) if pd.notna(row['crm_id']) else "")
                element_item = QTableWidgetItem(str(row['element']) if pd.notna(row['element']) else "")
                value_item = QTableWidgetItem(f"{row['value']:.6f}" if pd.notna(row['value']) else "")
                corrected_item = QTableWidgetItem(f"{row['corrected_value']:.6f}" if pd.notna(row['corrected_value']) else "")
                ref_item = QTableWidgetItem(f"{row['ref_value']:.6f}" if pd.notna(row['ref_value']) else "")
                
                # Calculate percentage difference
                percentage_diff = (abs(row['corrected_value'] - row['ref_value']) / row['ref_value'] * 100 
                                 if pd.notna(row['corrected_value']) and pd.notna(row['ref_value']) and row['ref_value'] != 0 
                                 else 0)
                diff_item = QTableWidgetItem(f"{percentage_diff:.2f}%")

                value_color = QColor('red') if row['out_no_blank'] else QColor('green')
                corrected_color = QColor('red') if row['out_with_blank'] else QColor('green')
                value_item.setForeground(value_color)
                corrected_item.setForeground(corrected_color)
                diff_item.setForeground(QColor('black'))

                self.table_widget.setItem(i, 0, crm_id_item)
                self.table_widget.setItem(i, 1, element_item)
                self.table_widget.setItem(i, 2, value_item)
                self.table_widget.setItem(i, 3, corrected_item)
                self.table_widget.setItem(i, 4, ref_item)
                self.table_widget.setItem(i, 5, diff_item)
        else:
            self.table_widget.setRowCount(0)

        self.table_widget.resizeColumnsToContents()
        self.table_widget.update()
        self.table_widget.repaint()

        self.export_button = QPushButton("Export to Excel")
        self.export_button.clicked.connect(self.export_to_excel)

        self.layout.addWidget(self.table_widget)
        self.layout.addWidget(self.export_button)
        self.setLayout(self.layout)

    def export_to_excel(self):
        if self.out_df is None or self.out_df.empty:
            QMessageBox.warning(self, "Warning", "No data to export")
            return

        fname, _ = QFileDialog.getSaveFileName(self, "Save Excel File", "", "Excel Files (*.xlsx)")
        if fname:
            try:
                # Prepare DataFrame for export
                export_df = self.out_df[['crm_id', 'element', 'value', 'corrected_value', 'ref_value']].copy()
                # Add percentage difference column (store as decimal, e.g., 0.5530 for 55.30%)
                export_df['Diff %'] = export_df.apply(
                    lambda row: (abs(row['corrected_value'] - row['ref_value']) / row['ref_value'] 
                                 if pd.notna(row['corrected_value']) and pd.notna(row['ref_value']) and row['ref_value'] != 0 
                                 else 0), 
                    axis=1
                )

                # Save to Excel with openpyxl
                with pd.ExcelWriter(fname, engine='openpyxl') as writer:
                    export_df.to_excel(writer, index=False, sheet_name='OutOfRange')
                    workbook = writer.book
                    worksheet = writer.sheets['OutOfRange']

                    # Define styles
                    red_font = Font(color='FF0000', bold=True)
                    green_font = Font(color='008000', bold=True)
                    black_font = Font(color='000000')
                    header_font = Font(bold=True, color='FFFFFF')
                    header_fill = PatternFill(start_color='0078D4', end_color='0078D4', fill_type='solid')
                    out_of_range_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
                    thin_border = Border(
                        left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin')
                    )
                    center_align = Alignment(horizontal='center', vertical='center')

                    # Apply header styles
                    for cell in worksheet[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_align
                        cell.border = thin_border

                    # Apply formatting to data rows
                    for row_idx, row in enumerate(self.out_df.itertuples(), start=2):
                        value_out = row.out_no_blank
                        corrected_out = row.out_with_blank

                        # Apply styles to cells
                        for col_idx, col_name in enumerate(['crm_id', 'element', 'value', 'corrected_value', 'ref_value', 'Diff %'], 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.border = thin_border
                            cell.alignment = center_align
                            
                            if col_name == 'value':
                                cell.font = red_font if value_out else green_font
                                cell.number_format = '0.000000'
                            elif col_name == 'corrected_value':
                                cell.font = red_font if corrected_out else green_font
                                cell.number_format = '0.000000'
                            elif col_name == 'ref_value':
                                cell.font = black_font
                                cell.number_format = '0.000000'
                            elif col_name == 'Diff %':
                                cell.font = black_font
                                cell.number_format = '0.00%'  # Format as percentage (0.5530 → 55.30%)
                            else:
                                cell.font = black_font

                            # Apply row background if out of range
                            if value_out or corrected_out:
                                cell.fill = out_of_range_fill

                    # Adjust column widths
                    for col in worksheet.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                max_length = max(max_length, len(str(cell.value)))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 30)  # Cap width at 30 for aesthetics
                        worksheet.column_dimensions[column].width = adjusted_width

                    # Add filters to the table
                    worksheet.auto_filter.ref = worksheet.dimensions

                QMessageBox.information(self, "Success", f"Data exported to {fname}")
                logger.info(f"Exported out-of-range data to {fname}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to export: {str(e)}")
                logger.error(f"Error exporting Excel: {str(e)}")

class OutOfRangeThread(QThread):
    out_of_range_data = pyqtSignal(pd.DataFrame)
    progress_updated = pyqtSignal(int)

    def __init__(self, db_path, file_name, percentage, ver_db_path):
        super().__init__()
        self.db_path = db_path
        self.file_name = file_name
        self.percentage = percentage
        self.ver_db_path = ver_db_path
        self.verification_cache = {}  # Initialize verification cache

    def run(self):
        try:
            self.progress_updated.emit(20)
            logger.info(f"Starting OutOfRangeThread for file: {self.file_name}")
            conn = sqlite3.connect(self.db_path)
            df = pd.read_sql_query("SELECT * FROM crm_data WHERE file_name = ?", conn, params=(self.file_name,))
            conn.close()

            crm_df = df[df['crm_id'] != 'BLANK'].copy()
            blank_df = df[df['crm_id'] == 'BLANK'].copy()
            crm_df['norm_crm_id'] = crm_df['crm_id'].apply(self.normalize_crm_id)

            out_df = pd.DataFrame()
            if crm_df.empty:
                logger.warning(f"No CRM data found for file {self.file_name}")
                self.out_of_range_data.emit(out_df)
                self.progress_updated.emit(100)
                return

            # Extract base elements (e.g., 'Ce' from 'Ce 140')
            crm_df['base_element'] = crm_df['element'].apply(lambda x: x.split()[0] if isinstance(x, str) and ' ' in x else x)
            unique_crms = crm_df['norm_crm_id'].unique()
            unique_base_elements = crm_df['base_element'].unique()

            self.progress_updated.emit(40)

            for crm_id in unique_crms:
                for base_element in unique_base_elements:
                    # Get all wavelengths for this base element and CRM ID
                    element_df = crm_df[(crm_df['norm_crm_id'] == crm_id) & (crm_df['base_element'] == base_element)]
                    if element_df.empty:
                        continue

                    ver_value = self.get_verification_value(crm_id, base_element)
                    if ver_value is None:
                        continue

                    lcl = ver_value * (1 - self.percentage / 100)
                    ucl = ver_value * (1 + self.percentage / 100)

                    # Calculate best wavelength based on corrected_value or value
                    best_row = None
                    min_diff = float('inf')

                    for _, row in element_df.iterrows():
                        value = row['value']
                        blank_value, corrected_value = self.select_best_blank(row, blank_df, ver_value)

                        # Use corrected_value if available, else use value
                        target_value = corrected_value if pd.notna(corrected_value) else value
                        diff = abs(target_value - ver_value)

                        if diff < min_diff:
                            min_diff = diff
                            best_row = {
                                'crm_id': row['crm_id'],
                                'element': row['element'],
                                'value': value,
                                'corrected_value': corrected_value if pd.notna(corrected_value) else pd.NA,
                                'ref_value': ver_value,
                                'out_no_blank': not (lcl <= value <= ucl),
                                'out_with_blank': not (lcl <= corrected_value <= ucl) if pd.notna(corrected_value) else not (lcl <= value <= ucl)
                            }

                    if best_row and (best_row['out_no_blank'] or best_row['out_with_blank']):
                        out_df = pd.concat([out_df, pd.DataFrame([best_row])], ignore_index=True)
                        logger.info(f"Added out-of-range record for CRM {crm_id}, Element {best_row['element']}: {best_row}")

            self.progress_updated.emit(100)
            self.out_of_range_data.emit(out_df)
        except Exception as e:
            logger.error(f"Error computing out of range for {self.file_name}: {str(e)}")
            self.out_of_range_data.emit(pd.DataFrame())
            self.progress_updated.emit(100)

    def normalize_crm_id(self, crm_id):
        """Extract numeric part from CRM ID (e.g., 'CRM 258b' → '258')."""
        if not isinstance(crm_id, str):
            return None
        crm_pattern = re.compile(r'^(?:\s*CRM\s*)?(\d{3})(?:\s*[a-zA-Z])?$', re.IGNORECASE)
        match = crm_pattern.match(crm_id.strip())
        if match:
            logger.debug(f"Normalized CRM ID: {crm_id} → {match.group(1)}")
            return match.group(1)
        logger.debug(f"Invalid CRM ID format: {crm_id}")
        return None

    def is_valid_crm_id(self, crm_id):
        """Check if CRM ID is valid."""
        norm = self.normalize_crm_id(crm_id)
        allowed_crms = ['258', '252', '906', '506', '233', '255', '263', '260']
        return norm in allowed_crms

    def get_verification_value(self, crm_id, element):
        cache_key = f"{crm_id}_{element}"
        if cache_key in self.verification_cache:
            logger.debug(f"Retrieved verification value from cache for {cache_key}: {self.verification_cache[cache_key]}")
            return self.verification_cache[cache_key]

        if not self.is_valid_crm_id(crm_id):
            logger.warning(f"Invalid CRM ID format: {crm_id}")
            self.verification_cache[cache_key] = None
            return None

        try:
            conn = sqlite3.connect(self.ver_db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [row[0] for row in cursor.fetchall()]
            table_name = "oreas_hs j" if re.match(r'(?i)oreas', crm_id) else "pivot_crm"
            if table_name not in tables:
                logger.error(f"Table {table_name} does not exist in database")
                conn.close()
                self.verification_cache[cache_key] = None
                return None

            cursor.execute(f"PRAGMA table_info({table_name})")
            cols = [x[1] for x in cursor.fetchall()]
            if 'CRM ID' not in cols:
                logger.error(f"Column 'CRM ID' not found in {table_name}")
                conn.close()
                self.verification_cache[cache_key] = None
                return None

            element_base = element.split()[0] if ' ' in element else element
            target_element = element if element in cols else element_base
            m = re.search(r'(?i)(?:CRM|OREAS)?\s*(\w+)(?:\s*par)?', crm_id)
            crm_id_part = m.group(1) if m else crm_id
            query = f"SELECT * FROM {table_name} WHERE [CRM ID] LIKE ?"
            cursor.execute(query, (f"%{crm_id_part}%",))
            crm_data = cursor.fetchall()

            if not crm_data:
                logger.warning(f"No CRM data found for {crm_id}")
                conn.close()
                self.verification_cache[cache_key] = None
                return None

            for row in crm_data:
                row_dict = {cols[i]: row[i] for i in range(len(cols))}
                label = str(row_dict['CRM ID']).strip().upper()
                if label.find(crm_id_part.upper()) != -1:
                    value = row_dict.get(target_element)
                    if value is not None and not pd.isna(value):
                        try:
                            value = float(value)
                            self.verification_cache[cache_key] = value
                            logger.debug(f"Verification value for CRM {crm_id}, Element {element}: {value}")
                            return value
                        except (ValueError, TypeError):
                            logger.warning(f"Invalid value for {target_element}: {value}")
                            continue

            logger.warning(f"No valid value for {target_element} in {table_name}")
            self.verification_cache[cache_key] = None
            return None
        except Exception as e:
            logger.error(f"Error querying verification database: {str(e)}")
            self.verification_cache[cache_key] = None
            return None
        finally:
            if 'conn' in locals():
                conn.close()

    def select_best_blank(self, crm_row, blank_df, ver_value):
        if blank_df.empty or ver_value is None:
            logger.debug(f"No blank correction applied: empty blank_df={blank_df.empty}, ver_value={ver_value}")
            return None, crm_row['value']

        relevant_blanks = blank_df[
            (blank_df['file_name'] == crm_row['file_name']) &
            (blank_df['folder_name'] == crm_row['folder_name']) &
            (blank_df['element'] == crm_row['element'])
        ]

        if relevant_blanks.empty:
            logger.debug(f"No relevant blanks found for CRM: file={crm_row['file_name']}, folder={crm_row['folder_name']}, element={crm_row['element']}")
            return None, crm_row['value']

        blank_valid_pattern = re.compile(r'^(?:CRM\s*)?(?:BLANK|BLNK|Blank|blnk|blank)(?:\s*[a-zA-Z]{1,2})?$', re.IGNORECASE)
        valid_blanks = relevant_blanks[relevant_blanks['solution_label'].apply(lambda x: bool(blank_valid_pattern.match(str(x).strip())))]

        if valid_blanks.empty:
            logger.debug(f"No valid blanks found for CRM row {crm_row['id']}")
            return None, crm_row['value']

        initial_diff = abs(crm_row['value'] - ver_value)
        best_blank_value = None
        best_diff = initial_diff
        corrected_value = crm_row['value']

        for _, blank_row in valid_blanks.iterrows():
            blank_value = blank_row['value']
            if pd.notna(blank_value):
                try:
                    corrected = crm_row['value'] - blank_value
                    new_diff = abs(ver_value - corrected)
                    logger.debug(f"Blank: solution_label={blank_row['solution_label']}, value={blank_value:.2f}, corrected={corrected:.2f}, new_diff={new_diff:.2f}, initial_diff={initial_diff:.2f}")
                    if new_diff < initial_diff:
                        best_diff = new_diff
                        best_blank_value = blank_value
                        corrected_value = corrected
                except (TypeError, ValueError) as e:
                    logger.warning(f"Invalid blank value {blank_value} for CRM row {crm_row['id']}: {str(e)}")
                    continue

        if best_blank_value is not None:
            logger.info(f"Selected blank value {best_blank_value:.2f} for CRM row {crm_row['id']}, corrected value={corrected_value:.2f}, diff={best_diff:.2f}")
        else:
            logger.debug(f"No valid blank value selected for CRM row {crm_row['id']}, using original value={crm_row['value']:.2f}")

        return best_blank_value, corrected_value
    
class CRMDataVisualizer(QMainWindow):
    def __init__(self):
        super().__init__()
        setTheme(Theme.LIGHT)
        self.setWindowTitle("CRM Data Visualizer")
        self.setGeometry(100, 100, 1400, 900)

        self.crm_df = pd.DataFrame()
        self.blank_df = pd.DataFrame()
        self.crm_db_path = "crm_blank.db"
        self.ver_db_path = "crm_data.db"
        self.filtered_crm_df_cache = None
        self.filtered_blank_df_cache = None
        self.plot_df_cache = None
        self.updating_filters = False
        self.verification_cache = {}
        self.plot_data_items = []
        self.logo_path = Path("logo.jpg")
        
        self.create_settings_table()

        self.main_widget = QWidget()
        self.setCentralWidget(self.main_widget)
        self.main_layout = QVBoxLayout(self.main_widget)
        self.main_layout.setSpacing(16)
        self.main_layout.setContentsMargins(20, 20, 20, 20)

        self.button_card = CardWidget()
        self.button_card.setStyleSheet("""
            CardWidget {
                background-color: #FFFFFF;
                border: 1px solid #E0E0E0;
                border-radius: 8px;
            }
        """)
        self.button_layout = QHBoxLayout()
        self.button_layout.setSpacing(12)
        self.button_layout.setContentsMargins(15, 10, 15, 10)
        self.import_button = PrimaryPushButton("Import File", self, FluentIcon.DOWNLOAD)
        self.export_button = PrimaryPushButton("Export Table", self, FluentIcon.SAVE)
        self.delete_button = PrimaryPushButton("Delete Files", self, FluentIcon.DELETE)
        self.edit_button = PrimaryPushButton("Edit Record", self, FluentIcon.EDIT)
        self.out_of_range_button = PrimaryPushButton("Out of Range", self, FluentIcon.SEARCH)
        self.save_button = PrimaryPushButton("Save Plot", self, FluentIcon.SAVE)
        self.reset_button = PrimaryPushButton("Reset Filters", self, FluentIcon.SYNC)
        self.button_layout.addWidget(self.import_button)
        self.button_layout.addWidget(self.export_button)
        self.button_layout.addWidget(self.delete_button)
        self.button_layout.addWidget(self.edit_button)
        self.button_layout.addWidget(self.out_of_range_button)
        self.button_layout.addWidget(self.save_button)
        self.button_layout.addWidget(self.reset_button)
        self.button_layout.addStretch()
        self.button_card.setLayout(self.button_layout)
        self.main_layout.addWidget(self.button_card)

        self.filter_logo_layout = QHBoxLayout()
        self.filter_logo_layout.setSpacing(16)

        self.filter_card = CardWidget()
        self.filter_card.setStyleSheet("""
            CardWidget {
                background-color: #FFFFFF;
                border: 1px solid #E0E0E0;
                border-radius: 8px;
            }
        """)
        self.filter_layout = QVBoxLayout()
        self.filter_layout.setSpacing(12)
        self.filter_layout.setContentsMargins(15, 15, 15, 15)

        self.filter_title = TitleLabel("Filter Controls")
        self.filter_title.setStyleSheet("""
            TitleLabel {
                color: #000000;
                font-size: 18px;
                font-weight: bold;
                padding: 8px 0;
            }
        """)
        self.filter_layout.addWidget(self.filter_title)

        self.controls_layout = QHBoxLayout()
        self.controls_layout.setSpacing(12)
        self.device_label = QLabel("Device:")
        self.device_combo = ComboBox()
        self.element_label = QLabel("Element:")
        self.element_combo = ComboBox()
        self.crm_label = QLabel("CRM ID:")
        self.crm_combo = ComboBox()
        self.from_date_label = QLabel("From Date:")
        self.from_date_edit = LineEdit()
        self.from_date_edit.setPlaceholderText("YYYY/MM/DD")
        self.from_date_edit.setFixedWidth(120)
        self.to_date_label = QLabel("To Date:")
        self.to_date_edit = LineEdit()
        self.to_date_edit.setPlaceholderText("YYYY/MM/DD")
        self.to_date_edit.setFixedWidth(120)
        self.percentage_label = QLabel("Control Range (%):")
        self.percentage_edit = LineEdit()
        self.percentage_edit.setPlaceholderText("%")
        self.percentage_edit.setFixedWidth(80)
        self.percentage_edit.setText("10")
        self.controls_layout.addWidget(self.device_label)
        self.controls_layout.addWidget(self.device_combo)
        self.controls_layout.addWidget(self.crm_label)
        self.controls_layout.addWidget(self.crm_combo)
        self.controls_layout.addWidget(self.element_label)
        self.controls_layout.addWidget(self.element_combo)
        self.controls_layout.addWidget(self.from_date_label)
        self.controls_layout.addWidget(self.from_date_edit)
        self.controls_layout.addWidget(self.to_date_label)
        self.controls_layout.addWidget(self.to_date_edit)
        self.controls_layout.addWidget(self.percentage_label)
        self.controls_layout.addWidget(self.percentage_edit)
        self.controls_layout.addStretch()
        self.filter_layout.addLayout(self.controls_layout)

        self.checkbox_layout = QVBoxLayout()
        self.checkbox_layout.setSpacing(8)
        self.best_wl_check = CheckBox("Select Best Wavelength")
        self.best_wl_check.setChecked(True)
        self.apply_blank_check = CheckBox("Apply Blank Correction")
        self.apply_blank_check.setChecked(False)
        self.checkbox_layout.addWidget(self.best_wl_check)
        self.checkbox_layout.addWidget(self.apply_blank_check)
        self.checkbox_layout.addStretch()
        self.filter_layout.addLayout(self.checkbox_layout)

        self.device_combo.setToolTip("Select a device to filter data")
        self.element_combo.setToolTip("Select an element to plot")
        self.crm_combo.setToolTip("Select a CRM ID to filter")
        self.from_date_edit.setToolTip("Enter start date in Jalali format (YYYY/MM/DD)")
        self.to_date_edit.setToolTip("Enter end date in Jalali format (YYYY/MM/DD)")
        self.percentage_edit.setToolTip("Enter control range percentage (e.g., 10 for ±10%)")
        self.best_wl_check.setToolTip("Select the best wavelength based on verification value")
        self.apply_blank_check.setToolTip("Subtract the best BLANK value from CRM data")

        self.device_combo.addItem("All Devices")
        self.element_combo.addItem("All Elements")
        self.crm_combo.addItem("All CRM IDs")

        self.filter_card.setLayout(self.filter_layout)

        self.logo_card = CardWidget()
        self.logo_card.setStyleSheet("""
            CardWidget {
                background-color: #FFFFFF;
                border: 1px solid #E0E0E0;
                border-radius: 8px;
            }
        """)
        self.logo_layout = QVBoxLayout()
        self.logo_layout.setContentsMargins(10, 10, 10, 10)
        self.logo_label = QLabel()
        self.logo_label.setFixedSize(140, 150)
        self.logo_layout.addWidget(self.logo_label)
        self.logo_card.setLayout(self.logo_layout)
        self.logo_card.setFixedWidth(120)

        self.filter_logo_layout.addWidget(self.filter_card, stretch=1)
        self.filter_logo_layout.addWidget(self.logo_card)
        self.main_layout.addLayout(self.filter_logo_layout)

        self.progress_bar = QProgressBar()
        self.progress_bar.setMaximum(100)
        self.progress_bar.setVisible(False)
        self.main_layout.addWidget(self.progress_bar)

        self.plot_widget = PlotWidget()
        self.plot_widget.setTitle("CRM Data Plot", color='#000000', size='14pt')
        self.plot_widget.setLabel('left', 'Value', color='#000000')
        self.plot_widget.setLabel('bottom', 'Observation', color='#000000')
        self.plot_widget.addLegend(offset=(10, 10))
        self.main_layout.addWidget(self.plot_widget, stretch=2)

        self.tooltip_label = QLabel("", self.plot_widget)
        self.tooltip_label.setStyleSheet("""
            background-color: #FFFFFF;
            color: #000000;
            border: 1px solid #0078D4;
            padding: 8px;
            border-radius: 4px;
            font-family: 'Segoe UI';
        """)
        self.tooltip_label.setVisible(False)
        self.tooltip_label.setFont(QFont("Segoe UI", 10))

        self.table_widget = QTableWidget()
        self.table_widget.setColumnCount(9)
        self.table_widget.setHorizontalHeaderLabels(["ID", "CRM ID", "Solution Label", "Element", "Value", "Blank Value", "File Name", "Date", "Ref Proximity %"])
        self.table_widget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table_widget.setSelectionMode(QTableWidget.SingleSelection)
        self.table_widget.setEditTriggers(QTableWidget.NoEditTriggers)
        self.main_layout.addWidget(self.table_widget, stretch=1)

        self.status_label = QLabel("Loading data...")
        self.main_layout.addWidget(self.status_label)

        self.device_combo.currentTextChanged.connect(self.on_device_or_crm_changed)
        self.element_combo.currentTextChanged.connect(self.on_filter_changed)
        self.crm_combo.currentTextChanged.connect(self.on_device_or_crm_changed)
        self.from_date_edit.textChanged.connect(self.on_filter_changed)
        self.to_date_edit.textChanged.connect(self.on_filter_changed)
        self.percentage_edit.textChanged.connect(self.on_filter_changed)
        self.best_wl_check.stateChanged.connect(self.on_filter_changed)
        self.apply_blank_check.stateChanged.connect(self.on_filter_changed)
        self.import_button.clicked.connect(self.import_file)
        self.export_button.clicked.connect(self.export_table)
        self.delete_button.clicked.connect(self.delete_files)
        self.edit_button.clicked.connect(self.edit_record)
        self.out_of_range_button.clicked.connect(self.show_out_of_range_dialog)
       
        # حذف دکمه → با تغییر element یا فیلترها خودکار پلات کن
        self.element_combo.currentTextChanged.connect(self.auto_plot)
        self.device_combo.currentTextChanged.connect(self.on_filter_changed)
        self.crm_combo.currentTextChanged.connect(self.on_filter_changed)
        self.from_date_edit.textChanged.connect(self.on_filter_changed)
        self.to_date_edit.textChanged.connect(self.on_filter_changed)
        self.percentage_edit.textChanged.connect(self.on_filter_changed)
        self.best_wl_check.stateChanged.connect(self.on_filter_changed)
        self.apply_blank_check.stateChanged.connect(self.on_filter_changed)

        self.save_button.clicked.connect(self.save_plot)
        self.reset_button.clicked.connect(self.reset_filters)
        # self.plot_widget.scene().sigMouseClicked.connect(self.on_mouse_clicked)  # Removed
        self.plot_widget.scene().sigMouseMoved.connect(self.on_mouse_moved)

        self.apply_styles()
        self.load_default_logo()

        logger.debug("Initializing CRMDataVisualizer")
        self.load_data_thread()
        
        self.create_settings_table()
        # بارگذاری تنظیمات ذخیره‌شده
        self.load_settings()

        # ذخیره خودکار هنگام بسته شدن
        self.setAttribute(Qt.WA_DeleteOnClose, False)  # اطمینان از اجرای closeEven

    def auto_plot(self):
        """خودکار پلات کردن با تغییر element و تنظیم خودکار محورها"""
        if self.updating_filters:
            return
        self.plot_data()
        # فعال‌سازی Auto Range
        self.plot_widget.enableAutoRange()

    def on_device_or_crm_changed(self):
        """وقتی Device یا CRM تغییر کرد، عناصر رو آپدیت کن"""
        if self.updating_filters:
            return
        self.save_settings()  # ذخیره فوری
        self.update_element_combo()  # بدون پارامتر

    def create_settings_table(self):
        """ایجاد جدول تنظیمات در دیتابیس اگر وجود نداشته باشد."""
        try:
            conn = sqlite3.connect(self.crm_db_path)
            cursor = conn.cursor()

            # بررسی وجود جدول settings
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='settings'")
            if cursor.fetchone():
                logger.debug("Settings table already exists")
            else:
                # ایجاد جدول جدید
                cursor.execute("""
                    CREATE TABLE settings (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        device TEXT,
                        element TEXT,
                        crm_id TEXT,
                        from_date TEXT,
                        to_date TEXT,
                        percentage TEXT,
                        best_wl_checked INTEGER,
                        apply_blank_checked INTEGER
                    )
                """)
                logger.debug("Settings table created")

                # درج ردیف پیش‌فرض
                cursor.execute("""
                    INSERT INTO settings (
                        device, element, crm_id, from_date, to_date, 
                        percentage, best_wl_checked, apply_blank_checked
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, ("All Devices", "All Elements", "All CRM IDs", "", "", "10", 1, 0))
                logger.info("Settings table initialized with default values")

            conn.commit()
        except Exception as e:
            logger.error(f"Error creating settings table: {str(e)}")
            raise
        finally:
            conn.close()

    def create_settings_table(self):
        """ایجاد جدول تنظیمات در دیتابیس اگر وجود نداشته باشد."""
        try:
            conn = sqlite3.connect(self.crm_db_path)
            cursor = conn.cursor()

            cursor.execute("""
                CREATE TABLE IF NOT EXISTS settings (
                    id INTEGER PRIMARY KEY CHECK (id = 1),
                    device TEXT,
                    element TEXT,
                    crm_id TEXT,
                    from_date TEXT,
                    to_date TEXT,
                    percentage TEXT,
                    best_wl_checked INTEGER,
                    apply_blank_checked INTEGER
                )
            """)

            # بررسی وجود ردیف
            cursor.execute("SELECT COUNT(*) FROM settings WHERE id = 1")
            if cursor.fetchone()[0] == 0:
                cursor.execute("""
                    INSERT INTO settings (
                        id, device, element, crm_id, from_date, to_date, 
                        percentage, best_wl_checked, apply_blank_checked
                    ) VALUES (1, ?, ?, ?, ?, ?, ?, ?, ?)
                """, ("mass", "", "", "", "", "10", 1, 0))

            conn.commit()
            logger.info("Settings table initialized")
        except Exception as e:
            logger.error(f"Error creating settings table: {str(e)}")
        finally:
            conn.close()


    def save_settings(self):
        """ذخیره تنظیمات فعلی در دیتابیس."""
        try:
            conn = sqlite3.connect(self.crm_db_path)
            cursor = conn.cursor()

            # اعتبارسنجی تاریخ
            from_date = self.from_date_edit.text().strip()
            if from_date and not validate_jalali_date(from_date):
                from_date = ""

            to_date = self.to_date_edit.text().strip()
            if to_date and not validate_jalali_date(to_date):
                to_date = ""

            # اعتبارسنجی درصد
            percentage = self.percentage_edit.text().strip()
            if not validate_percentage(percentage):
                percentage = "10"

            cursor.execute("""
                UPDATE settings SET
                    device = ?, element = ?, crm_id = ?, from_date = ?, to_date = ?,
                    percentage = ?, best_wl_checked = ?, apply_blank_checked = ?
                WHERE id = 1
            """, (
                self.device_combo.currentText(),
                self.element_combo.currentText(),
                self.crm_combo.currentText(),
                from_date,
                to_date,
                percentage,
                1 if self.best_wl_check.isChecked() else 0,
                1 if self.apply_blank_check.isChecked() else 0
            ))
            conn.commit()
            logger.debug("Settings saved to database")
        except Exception as e:
            logger.error(f"Error saving settings: {str(e)}")
        finally:
            conn.close()


    def load_settings(self):
        try:
            conn = sqlite3.connect(self.crm_db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM settings WHERE id = 1")
            row = cursor.fetchone()
            conn.close()

            if not row:
                logger.warning("No settings found")
                return

            _, device, element, crm_id, from_date, to_date, percentage, best_wl, apply_blank = row
            self._pending_settings = {
                'device': device,
                'element': element,
                'crm_id': crm_id,
                'from_date': from_date,
                'to_date': to_date,
                'percentage': percentage,
                'best_wl': int(best_wl),
                'apply_blank': int(apply_blank)
            }
            logger.info(f"Settings loaded: {self._pending_settings}")

        except Exception as e:
            logger.error(f"Error loading settings: {str(e)}")
            self._pending_settings = None

    def get_db_path(self, name):
        return name
        # return Path(__file__).parent / name

    def load_default_logo(self):
        if self.logo_path.exists():
            pixmap = QPixmap(str(self.logo_path))
            self.logo_label.setPixmap(pixmap.scaled(100, 120))
            logger.info(f"Default logo loaded: {self.logo_path}")
        else:
            logger.warning(f"Default logo not found at: {self.logo_path}")

    def load_data_thread(self):
        self.progress_bar.setVisible(True)
        self.loader_thread = DataLoaderThread(self.crm_db_path)
        self.loader_thread.data_loaded.connect(self.on_data_loaded)
        self.loader_thread.error_occurred.connect(self.on_data_error)
        self.loader_thread.progress_updated.connect(self.progress_bar.setValue)
        self.loader_thread.finished.connect(lambda: self.progress_bar.setVisible(False))
        self.loader_thread.start()

    def import_file(self):
        fname, _ = QFileDialog.getOpenFileName(self, "Import File", "", "CSV or REP Files (*.csv *.rep)")
        if fname:
            dialog = DeviceSelectionDialog(self)
            if dialog.exec_() == QDialog.Accepted:
                selected_device = dialog.get_device()
                logger.debug(f"Selected device for import: {selected_device}")
                self.progress_bar.setVisible(True)
                self.import_thread = ImportFileThread(fname, self.crm_db_path, selected_device)
                self.import_thread.import_completed.connect(self.on_import_completed)
                self.import_thread.error_occurred.connect(self.on_data_error)
                self.import_thread.progress_updated.connect(self.progress_bar.setValue)
                self.import_thread.finished.connect(lambda: self.progress_bar.setVisible(False))
                self.import_thread.start()
            else:
                logger.debug("Import canceled by user")

    def delete_files(self):
        all_df = pd.concat([self.crm_df, self.blank_df])
        file_names = all_df['file_name'].unique()
        if len(file_names) == 0:
            QMessageBox.warning(self, "Warning", "No files to delete")
            return
        
        dialog = DeleteFilesDialog(self, file_names, self.crm_db_path)
        if dialog.exec_() == QDialog.Accepted:
            selected_files = dialog.get_selected_files()
            if not selected_files:
                QMessageBox.warning(self, "Warning", "No files selected for deletion")
                return
            
            total_records = sum(dialog.record_counts.get(f, 0) for f in selected_files)
            confirm = QMessageBox.question(
                self, "Confirm Delete",
                f"Are you sure you want to delete {len(selected_files)} files with {total_records} records?",
                QMessageBox.Yes | QMessageBox.No
            )
            if confirm != QMessageBox.Yes:
                logger.debug("Deletion canceled by user")
                return
            
            self.progress_bar.setVisible(True)
            self.delete_thread = DeleteFilesThread(self.crm_db_path, selected_files)
            self.delete_thread.delete_completed.connect(self.on_delete_completed)
            self.delete_thread.error_occurred.connect(self.on_data_error)
            self.delete_thread.progress_updated.connect(self.progress_bar.setValue)
            self.delete_thread.finished.connect(lambda: self.progress_bar.setVisible(False))
            self.delete_thread.start()

    def on_delete_completed(self):
        self.load_data_thread()
        self.status_label.setText("Files deleted successfully")
        logger.info("Files deleted successfully")

    def on_import_completed(self, df):
        self.load_data_thread()
        self.status_label.setText(f"Imported {len(df)} records successfully")
        logger.info(f"Imported {len(df)} records successfully")

    def on_data_loaded(self, crm_df, blank_df):
        self.crm_df = crm_df
        self.blank_df = blank_df
        logger.info(f"Loaded {len(crm_df)} CRM records and {len(blank_df)} BLANK records")
        
        self.populate_filters()
        self.status_label.setText("Data loaded successfully")

        # بعد از پر شدن ComboBoxها، تنظیمات رو بازیابی کن
        if hasattr(self, '_loaded_settings'):
            QTimer.singleShot(100, self.restore_filters_after_load)

    def on_data_error(self, error_message):
        self.crm_df = pd.DataFrame()
        self.blank_df = pd.DataFrame()
        self.status_label.setText(error_message)
        logger.error(error_message)
        QMessageBox.critical(self, "Error", error_message)
        self.populate_filters()

    def closeEvent(self, event):
        """ذخیره تنظیمات هنگام بسته شدن پنجره."""
        try:
            self.save_settings()
            logger.info("Settings saved on application close")
        except Exception as e:
            logger.error(f"Error saving settings on close: {str(e)}")
        finally:
            event.accept()

    def on_filter_changed(self):
        """فیلترها تغییر کردند → ذخیره فوری + آپدیت"""
        if self.updating_filters:
            return

        self.updating_filters = True
        try:
            # ذخیره فوری تنظیمات
            self.save_settings()

            # --- استخراج فیلترها ---
            from_date = None
            if validate_jalali_date(self.from_date_edit.text()):
                y, m, d = map(int, self.from_date_edit.text().split('/'))
                from_date = JalaliDate(y, m, d)

            to_date = None
            if validate_jalali_date(self.to_date_edit.text()):
                y, m, d = map(int, self.to_date_edit.text().split('/'))
                to_date = JalaliDate(y, m, d)

            filters = {
                'device': self.device_combo.currentText(),
                'element': self.element_combo.currentText(),
                'crm': self.crm_combo.currentText(),
                'from_date': from_date,
                'to_date': to_date
            }

            # --- اجرای فیلتر در ترد جدا ---
            self.progress_bar.setVisible(True)
            self.filter_thread = FilterThread(self.crm_df, self.blank_df, filters)
            self.filter_thread.filtered_data.connect(self.on_filtered_data)
            self.filter_thread.progress_updated.connect(self.progress_bar.setValue)
            self.filter_thread.finished.connect(lambda: self.progress_bar.setVisible(False))
            self.filter_thread.start()

        except Exception as e:
            logger.error(f"Error in on_filter_changed: {str(e)}")
        finally:
            self.updating_filters = False

    def is_valid_crm_id(self, crm_id):
        norm = normalize_crm_id(crm_id)
        allowed_crms = ['258', '252', '906', '506', '233', '255', '263', '260']
        return norm in allowed_crms

    def extract_device_name(self, folder_name):
        if not folder_name or not isinstance(folder_name, str):
            return None
        allowed_devices = {'mass', 'oes 4ac', 'oes fire'}
        normalized_name = folder_name.strip().lower()
        if normalized_name in allowed_devices:
            return normalized_name
        return None
    
    def restore_filters_after_load(self):
        """بازیابی مقادیر فیلتر بعد از لود داده و پر شدن ComboBoxها"""
        if not hasattr(self, '_loaded_settings'):
            return

        self.updating_filters = True
        try:
            device, element, crm_id, from_date, to_date, percentage, best_wl, apply_blank = self._loaded_settings

            # Device
            if device and self.device_combo.findText(device) != -1:
                self.device_combo.setCurrentText(device)

            # CRM
            if crm_id and self.crm_combo.findText(crm_id) != -1:
                self.crm_combo.setCurrentText(crm_id)

            # تاریخ
            if validate_jalali_date(from_date):
                self.from_date_edit.setText(from_date)
            if validate_jalali_date(to_date):
                self.to_date_edit.setText(to_date)

            # درصد
            if validate_percentage(percentage):
                self.percentage_edit.setText(percentage)

            # چک‌باکس‌ها
            self.best_wl_check.setChecked(bool(best_wl))
            self.apply_blank_check.setChecked(bool(apply_blank))

            # حالا element رو آپدیت کن (بعد از device و crm)
            QTimer.singleShot(100, lambda: self.update_element_combo(element))

        except Exception as e:
            logger.error(f"Error restoring filters: {str(e)}")
        finally:
            self.updating_filters = False
            # بعد از همه، فیلترها رو اعمال کن
            QTimer.singleShot(200, self.update_filters)

    def restore_element(self, target_element):
        self.update_element_combo()  # اول آپدیت کن
        QTimer.singleShot(50, lambda: self.set_element_safely(target_element))

    def set_element_safely(self, target_element):
        if not target_element:
            return
        items = [self.element_combo.itemText(i) for i in range(self.element_combo.count())]
        if target_element in items:
            self.element_combo.setCurrentText(target_element)
            logger.info(f"Element restored: {target_element}")
        else:
            logger.warning(f"Element {target_element} not found in current filter")
    def apply_pending_settings(self):
        if not hasattr(self, '_pending_settings') or not self._pending_settings:
            return

        settings = self._pending_settings
        self.updating_filters = True
        try:
            # 1. Device
            if settings['device'] and self.device_combo.findText(settings['device']) != -1:
                self.device_combo.setCurrentText(settings['device'])

            # 2. CRM
            if settings['crm_id'] and self.crm_combo.findText(settings['crm_id']) != -1:
                self.crm_combo.setCurrentText(settings['crm_id'])

            # 3. تاریخ و درصد
            if validate_jalali_date(settings['from_date']):
                self.from_date_edit.setText(settings['from_date'])
            if validate_jalali_date(settings['to_date']):
                self.to_date_edit.setText(settings['to_date'])
            if validate_percentage(settings['percentage']):
                self.percentage_edit.setText(settings['percentage'])

            # 4. چک‌باکس‌ها
            self.best_wl_check.setChecked(bool(settings['best_wl']))
            self.apply_blank_check.setChecked(bool(settings['apply_blank']))

            # 5. Element — بعد از device و crm
            QTimer.singleShot(150, lambda: self.restore_element(settings['element']))

        except Exception as e:
            logger.error(f"Error applying settings: {str(e)}")
        finally:
            self.updating_filters = False
            QTimer.singleShot(300, self.update_filters)  # بعد از همه، فیلتر کن
            
    def on_data_loaded(self, crm_df, blank_df):
        self.crm_df = crm_df
        self.blank_df = blank_df
        self.populate_filters()
        self.status_label.setText("Data loaded")
        # apply_pending_settings در populate_filters فراخوانی میشه
    def populate_filters(self):
        if self.crm_df.empty and self.blank_df.empty:
            return

        self.updating_filters = True
        try:
            # --- Device ---
            self.device_combo.blockSignals(True)
            self.device_combo.clear()
            self.device_combo.addItems(['mass', 'oes 4ac', 'oes fire'])
            self.device_combo.blockSignals(False)

            # --- CRM ---
            self.crm_combo.blockSignals(True)
            self.crm_combo.clear()
            crms = sorted(self.crm_df['norm_crm_id'].dropna().unique())
            self.crm_combo.addItems(crms)
            self.crm_combo.blockSignals(False)

            # --- Element ---
            self.element_combo.blockSignals(True)
            self.element_combo.clear()
            self.element_combo.blockSignals(False)

        except Exception as e:
            logger.error(f"Error in populate_filters: {str(e)}")
        finally:
            self.updating_filters = False

        # بعد از پر شدن ComboBoxها، تنظیمات رو اعمال کن
        QTimer.singleShot(100, self.apply_pending_settings)
        
    def update_filters(self):
        if self.updating_filters:
            return
        self.updating_filters = True

        try:
            if self.crm_df.empty and self.blank_df.empty:
                self.table_widget.setRowCount(0)
                self.status_label.setText("No data available")
                logger.warning("No data available for filtering")
                return

            from_date = None
            if validate_jalali_date(self.from_date_edit.text()):
                y, m, d = map(int, self.from_date_edit.text().split('/'))
                from_date = JalaliDate(y, m, d)

            to_date = None
            if validate_jalali_date(self.to_date_edit.text()):
                y, m, d = map(int, self.to_date_edit.text().split('/'))
                to_date = JalaliDate(y, m, d)

            filters = {
                'device': self.device_combo.currentText(),
                'element': self.element_combo.currentText(),
                'crm': self.crm_combo.currentText(),
                'from_date': from_date,
                'to_date': to_date
            }
            # logger.debug(f"Updating filters: {filters}")

            # فقط در صورتی که فیلترها تغییر کرده باشند، ذخیره کن
            current_settings = {
                'device': self.device_combo.currentText(),
                'element': self.element_combo.currentText(),
                'crm_id': self.crm_combo.currentText(),
                'from_date': self.from_date_edit.text(),
                'to_date': self.to_date_edit.text(),
                'percentage': self.percentage_edit.text(),
                'best_wl_checked': 1 if self.best_wl_check.isChecked() else 0,
                'apply_blank_checked': 1 if self.apply_blank_check.isChecked() else 0
            }

            # بررسی تغییرات با تنظیمات قبلی
            conn = sqlite3.connect(self.crm_db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM settings WHERE id = 1")
            saved_settings = cursor.fetchone()
            conn.close()

            if saved_settings:
                saved_settings_dict = {
                    'device': saved_settings[1],
                    'element': saved_settings[2],
                    'crm_id': saved_settings[3],
                    'from_date': saved_settings[4],
                    'to_date': saved_settings[5],
                    'percentage': saved_settings[6],
                    'best_wl_checked': saved_settings[7],
                    'apply_blank_checked': saved_settings[8]
                }
                if current_settings != saved_settings_dict:
                    self.save_settings()

            self.progress_bar.setVisible(True)
            self.filter_thread = FilterThread(self.crm_df, self.blank_df, filters)
            self.filter_thread.filtered_data.connect(self.on_filtered_data)
            self.filter_thread.progress_updated.connect(self.progress_bar.setValue)
            self.filter_thread.finished.connect(lambda: self.progress_bar.setVisible(False))
            self.filter_thread.start()

        finally:
            self.updating_filters = False

    def on_filtered_data(self, filtered_crm_df, filtered_blank_df):
        self.filtered_crm_df_cache = filtered_crm_df
        self.filtered_blank_df_cache = filtered_blank_df
        QApplication.processEvents()
        self.status_label.setText(f"Loaded {len(filtered_crm_df)} CRM records, {len(filtered_blank_df)} BLANK records")
        logger.info(f"Filtered {len(filtered_crm_df)} CRM records and {len(filtered_blank_df)} BLANK records")

        # خودکار پلات کردن بعد از فیلتر
        self.auto_plot()

    def update_element_combo(self, target_element=None):
        """آپدیت عناصر بر اساس Device و CRM — با هدف بازیابی مقدار قبلی"""
        if self.updating_filters:
            return

        self.updating_filters = True
        try:
            current_device = self.device_combo.currentText()
            current_crm = self.crm_combo.currentText()

            self.element_combo.blockSignals(True)
            previous_element = self.element_combo.currentText()
            self.element_combo.clear()

            # فیلتر داده
            mask = pd.Series([True] * len(self.crm_df), index=self.crm_df.index)
            if current_device:
                mask &= self.crm_df['folder_name'].str.contains(current_device, case=False, na=False)
            if current_crm:
                mask &= (self.crm_df['norm_crm_id'] == current_crm)

            filtered = self.crm_df[mask]
            if not filtered.empty:
                elements = sorted({
                    el.split()[0] for el in filtered['element'].dropna().unique()
                    if isinstance(el, str) and ' ' in el
                })
                self.element_combo.addItems(elements)

            # بازیابی مقدار هدف (از تنظیمات ذخیره‌شده)
            if target_element and target_element in elements:
                self.element_combo.setCurrentText(target_element)
            elif previous_element in [self.element_combo.itemText(i) for i in range(self.element_combo.count())]:
                self.element_combo.setCurrentText(previous_element)
            elif self.element_combo.count() > 0:
                self.element_combo.setCurrentIndex(0)

        except Exception as e:
            logger.error(f"Error in update_element_combo: {str(e)}")
        finally:
            self.element_combo.blockSignals(False)
            self.updating_filters = False

        # بعد از آپدیت عنصر، پلات کن
        QTimer.singleShot(50, self.auto_plot)


    def update_table(self, crm_df, blank_df):
        self.table_widget.blockSignals(True)
        combined_df = self.plot_df_cache if self.plot_df_cache is not None else pd.DataFrame()
        if combined_df.empty:
            logger.debug("No plotted data to display in table")
            self.table_widget.setRowCount(0)
            self.table_widget.blockSignals(False)
            return

        for col in ['id', 'crm_id', 'solution_label', 'element', 'value', 'blank_value', 'file_name', 'date']:
            if col not in combined_df.columns:
                combined_df[col] = pd.NA
        combined_df = combined_df.sort_values(['date', 'crm_id', 'element'])

        current_element = self.element_combo.currentText()
        current_crm = self.crm_combo.currentText()
        ver_value = None
        if current_element != "All Elements" and current_crm != "All CRM IDs":
            ver_value = self.get_verification_value(current_crm, current_element)

        combined_df['ref_proximity'] = pd.NA
        if ver_value is not None:
            if self.apply_blank_check.isChecked():
                combined_df['ref_proximity'] = abs(combined_df['value'] - ver_value) / ver_value * 100 if 'value' in combined_df else pd.NA
            else:
                combined_df['ref_proximity'] = abs(combined_df['value'] - ver_value) / ver_value * 100 if 'value' in combined_df else pd.NA

        self.table_widget.setRowCount(len(combined_df))
        for i, row in combined_df.iterrows():
            QApplication.processEvents()
            self.table_widget.setItem(i, 0, QTableWidgetItem(str(row['id']) if pd.notna(row['id']) else ""))
            self.table_widget.setItem(i, 1, QTableWidgetItem(str(row['crm_id'])))
            self.table_widget.setItem(i, 2, QTableWidgetItem(str(row['solution_label'])))
            self.table_widget.setItem(i, 3, QTableWidgetItem(str(row['element'])))
            self.table_widget.setItem(i, 4, QTableWidgetItem(f"{row['value']:.2f}" if pd.notna(row['value']) else ""))
            self.table_widget.setItem(i, 5, QTableWidgetItem(f"{row['blank_value']:.2f}" if pd.notna(row['blank_value']) else ""))
            self.table_widget.setItem(i, 6, QTableWidgetItem(str(row['file_name'])))
            self.table_widget.setItem(i, 7, QTableWidgetItem(str(row['date']) if pd.notna(row['date']) else ""))
            self.table_widget.setItem(i, 8, QTableWidgetItem(f"{row['ref_proximity']:.2f}%" if pd.notna(row['ref_proximity']) else ""))
        
        self.status_label.setText(f"Loaded {len(combined_df)} plotted CRM records")
        logger.info(f"Updated table with {len(combined_df)} plotted records")
        self.table_widget.blockSignals(False)

    def export_table(self):
        if self.plot_df_cache is None or self.plot_df_cache.empty:
            QMessageBox.warning(self, "Warning", "No data to export")
            return
        fname, _ = QFileDialog.getSaveFileName(self, "Save CSV", "", "CSV (*.csv)")
        if fname:
            try:
                self.plot_df_cache.to_csv(fname, index=False, encoding='utf-8')
                self.status_label.setText("Table exported successfully")
                logger.info(f"Table exported to {fname}")
            except Exception as e:
                logger.error(f"Error exporting table: {str(e)}")
                QMessageBox.critical(self, "Error", f"Failed to export table: {str(e)}")

    def get_verification_value(self, crm_id, element):
        cache_key = f"{crm_id}_{element}"
        if cache_key in self.verification_cache:
            logger.debug(f"Retrieved verification value from cache for {cache_key}: {self.verification_cache[cache_key]}")
            return self.verification_cache[cache_key]

        if not self.is_valid_crm_id(crm_id):
            logger.warning(f"Invalid CRM ID format: {crm_id}")
            self.verification_cache[cache_key] = None
            return None

        try:
            conn = sqlite3.connect(self.ver_db_path)
            cursor = conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables = [row[0] for row in cursor.fetchall()]
            table_name = "oreas_hs j" if re.match(r'(?i)oreas', crm_id) else "pivot_crm"
            if table_name not in tables:
                logger.error(f"Table {table_name} does not exist in database")
                conn.close()
                QMessageBox.critical(self, "Error", f"Table {table_name} does not exist")
                self.verification_cache[cache_key] = None
                return None

            cursor.execute(f"PRAGMA table_info({table_name})")
            cols = [x[1] for x in cursor.fetchall()]
            if 'CRM ID' not in cols:
                logger.error(f"Column 'CRM ID' not found in {table_name}")
                conn.close()
                QMessageBox.critical(self, "Error", f"Column 'CRM ID' not found")
                self.verification_cache[cache_key] = None
                return None

            element_base = element.split()[0] if ' ' in element else element
            target_element = element if element in cols else element_base
            m = re.search(r'(?i)(?:CRM|OREAS)?\s*(\w+)(?:\s*par)?', crm_id)
            crm_id_part = m.group(1) if m else crm_id
            query = f"SELECT * FROM {table_name} WHERE [CRM ID] LIKE ?"
            cursor.execute(query, (f"%{crm_id_part}%",))
            crm_data = cursor.fetchall()

            if not crm_data:
                logger.warning(f"No CRM data found for {crm_id}")
                conn.close()
                self.verification_cache[cache_key] = None
                return None

            for row in crm_data:
                row_dict = {cols[i]: row[i] for i in range(len(cols))}
                label = str(row_dict['CRM ID']).strip().upper()
                if label.find(crm_id_part.upper()) != -1:
                    value = row_dict.get(target_element)
                    if value is not None and not pd.isna(value):
                        try:
                            value = float(value)
                            self.verification_cache[cache_key] = value
                            logger.debug(f"Verification value for CRM {crm_id}, Element {element}: {value}")
                            return value
                        except (ValueError, TypeError):
                            logger.warning(f"Invalid value for {target_element}: {value}")
                            continue

            logger.warning(f"No valid value for {target_element} in {table_name}")
            self.verification_cache[cache_key] = None
            return None
        except Exception as e:
            logger.error(f"Error querying database: {str(e)}")
            QMessageBox.critical(self, "Error", f"Error querying database: {str(e)}")
            self.verification_cache[cache_key] = None
            return None
        finally:
            if 'conn' in locals():
                conn.close()

    def select_best_blank(self, crm_row, blank_df, ver_value):
        if blank_df.empty or ver_value is None:
            logger.debug(f"No blank correction applied: empty blank_df={blank_df.empty}, ver_value={ver_value}")
            return None, crm_row['value']
        
        relevant_blanks = blank_df[
            (blank_df['file_name'] == crm_row['file_name']) &
            (blank_df['folder_name'] == crm_row['folder_name']) &
            (blank_df['element'] == crm_row['element'])
        ]
        
        if relevant_blanks.empty:
            logger.debug(f"No relevant blanks found for CRM: file={crm_row['file_name']}, folder={crm_row['folder_name']}, element={crm_row['element']}")
            return None, crm_row['value']
        
        # Valid BLANK pattern: without 'par', usually with 1-2 letters
        blank_valid_pattern = re.compile(r'^(?:CRM\s*)?(?:BLANK|BLNK|Blank|blnk|blank)(?:\s*[a-zA-Z]{1,2})?$', re.IGNORECASE)
        
        valid_blanks = relevant_blanks[relevant_blanks['solution_label'].apply(lambda x: bool(blank_valid_pattern.match(str(x).strip())))]
        print(relevant_blanks,'valid blank :',valid_blanks)
        if valid_blanks.empty:
            logger.debug(f"No valid blanks found for CRM row {crm_row['id']}")
            return None, crm_row['value']
        
        initial_diff = abs(crm_row['value'] - ver_value)
        best_blank_value = None
        best_diff = initial_diff
        corrected_value = crm_row['value']
        
        for _, blank_row in valid_blanks.iterrows():
            blank_value = blank_row['value']
            if pd.notna(blank_value):
                try:
                    corrected = crm_row['value'] - blank_value
                    new_diff = abs(ver_value - corrected)
                    # print(corrected,ver_value,blank_value)
                    # logger.debug(f"Blank: solution_label={blank_row['solution_label']}, value={blank_value}, corrected={corrected}, new_diff={new_diff}, initial_diff={initial_diff}")
                    if new_diff < initial_diff:
                        best_diff = new_diff
                        best_blank_value = blank_value
                        corrected_value = corrected
                except (TypeError, ValueError) as e:
                    logger.warning(f"Invalid blank value {blank_value} for CRM row {crm_row['id']}: {str(e)}")
                    continue
        
        if best_blank_value is not None:
            logger.info(f"Selected blank value {best_blank_value} for CRM row {crm_row['id']}, corrected value={corrected_value}, diff={best_diff}")
        else:
            logger.debug(f"No valid blank value selected for CRM row {crm_row['id']}, using original value={crm_row['value']}")
        
        return best_blank_value, corrected_value

    def plot_data(self):
        self.plot_widget.clear()
        self.plot_data_items = []
        filtered_crm_df = self.filtered_crm_df_cache if self.filtered_crm_df_cache is not None else self.crm_df
        filtered_blank_df = self.filtered_blank_df_cache if self.filtered_blank_df_cache is not None else self.blank_df

        if filtered_crm_df.empty and filtered_blank_df.empty:
            self.status_label.setText("No data to plot")
            logger.info("No data to plot due to empty filtered dataframes")
            self.plot_df_cache = pd.DataFrame()
            self.update_table(pd.DataFrame(), pd.DataFrame())
            return

        percentage = 10.0
        if validate_percentage(self.percentage_edit.text()):
            percentage = float(self.percentage_edit.text())
        else:
            logger.warning(f"Invalid percentage value: {self.percentage_edit.text()}, using default 10%")
            self.percentage_edit.setText("10")

        filtered_crm_df = filtered_crm_df.sort_values('date')
        colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEEAD', '#D4A5A5', '#9B59B6']
        plot_df = pd.DataFrame()
        plotted_records = 0

        crm_ids = [self.crm_combo.currentText()] if self.crm_combo.currentText() != "All CRM IDs" else filtered_crm_df['norm_crm_id'].unique()
        logger.debug(f"Plotting for CRM IDs: {crm_ids}")

        for idx, crm_id in enumerate(crm_ids):
            crm_df = filtered_crm_df[filtered_crm_df['norm_crm_id'] == crm_id]
            if crm_df.empty:
                logger.debug(f"No data for CRM ID {crm_id}")
                continue

            current_element = self.element_combo.currentText()
            ver_value = self.get_verification_value(crm_id, current_element) if current_element != "All Elements" else None

            if current_element != "All Elements" and self.best_wl_check.isChecked() and ver_value is not None:
                def select_best(group):
                    group['diff'] = abs(group['value'] - ver_value)
                    return group.loc[group['diff'].idxmin()]
                crm_df = crm_df.groupby(['year', 'month', 'day']).apply(select_best).reset_index(drop=True)

            original_df = crm_df.copy()
            if self.apply_blank_check.isChecked() and current_element != "All Elements" and self.crm_combo.currentText() != "All CRM IDs" and ver_value is not None:
                crm_df = crm_df.copy()
                crm_df['original_value'] = crm_df['value']
                crm_df['blank_value'] = pd.NA
                for i, row in crm_df.iterrows():
                    blank_value, corrected_value = self.select_best_blank(row, filtered_blank_df, ver_value)
                    crm_df.at[i, 'value'] = corrected_value
                    crm_df.at[i, 'blank_value'] = blank_value

            indices = np.arange(len(crm_df))
            values = crm_df['value'].values
            original_values = original_df['value'].values if self.apply_blank_check.isChecked() else None
            date_labels = [d for d in crm_df['date']]
            logger.debug(f"CRM {crm_id}: {len(indices)} points, values range: {min(values, default=0):.2f} - {max(values, default=0):.2f}")

            # Adjust x_range for single point
            min_x = 0
            max_x = max(indices, default=0)
            if len(indices) == 1:
                max_x = 1

            x_range = [min_x, max_x]

            pen = mkPen(color=colors[idx % len(colors)], width=2)
            plot_item = self.plot_widget.plot(indices, values, pen=pen, symbol='o', symbolSize=8, name=f"CRM {crm_id} (Corrected)" if self.apply_blank_check.isChecked() else f"CRM {crm_id}")
            self.plot_data_items.append((plot_item, crm_df, indices, date_labels))

            if self.apply_blank_check.isChecked() and original_values is not None:
                original_pen = mkPen(color=colors[(idx + 1) % len(colors)], width=1, style=Qt.DashLine)
                original_plot_item = self.plot_widget.plot(indices, original_values, pen=original_pen, symbol='x', symbolSize=6, name=f"CRM {crm_id} (Original)")
                self.plot_data_items.append((original_plot_item, original_df, indices, date_labels))

            logger.debug(f"Plotted {len(crm_df)} points for CRM ID {crm_id}")
            plotted_records += len(crm_df)

            if current_element != "All Elements" and self.crm_combo.currentText() != "All CRM IDs":
                ver_value = self.get_verification_value(crm_id, current_element)
                if ver_value is not None and not pd.isna(ver_value):
                    delta = ver_value * (percentage / 100) / 3
                    self.plot_widget.plot(x_range, [ver_value * (1 - percentage / 100)] * 2, pen=mkPen('#FF6B6B', width=2, style=Qt.DotLine), name="LCL")
                    self.plot_widget.plot(x_range, [ver_value - 2 * delta] * 2, pen=mkPen('#4ECDC4', width=1, style=Qt.DotLine), name="-2LS")
                    self.plot_widget.plot(x_range, [ver_value - delta] * 2, pen=mkPen('#4ECDC4', width=1, style=Qt.DotLine), name="-1LS")
                    self.plot_widget.plot(x_range, [ver_value] * 2, pen=mkPen('#000000', width=3, style=Qt.DashLine), name=f"Ref Value ({ver_value:.3f})")
                    self.plot_widget.plot(x_range, [ver_value + delta] * 2, pen=mkPen('#45B7D1', width=1, style=Qt.DotLine), name="1LS")
                    self.plot_widget.plot(x_range, [ver_value + 2 * delta] * 2, pen=mkPen('#45B7D1', width=1, style=Qt.DotLine), name="2LS")
                    self.plot_widget.plot(x_range, [ver_value * (1 + percentage / 100)] * 2, pen=mkPen('#FF6B6B', width=2, style=Qt.DotLine), name="UCL")
                    logger.info(f"Plotted control lines for CRM {crm_id}, Element {current_element}")

            plot_df = pd.concat([plot_df, crm_df], ignore_index=True)
        
        self.plot_df_cache = plot_df
        if plotted_records == 0:
            self.status_label.setText("No data to plot")
            logger.info("No data plotted")
        else:
            self.status_label.setText(f"Plotted {plotted_records} records")
            logger.info(f"Plotted {plotted_records} records")
        self.plot_widget.enableAutoRange()  # فعال‌سازی Auto Range
        self.update_table(plot_df, pd.DataFrame())

    def show_out_of_range_dialog(self):
        all_df = pd.concat([self.crm_df, self.blank_df])
        file_names = all_df['file_name'].unique()
        if len(file_names) == 0:
            QMessageBox.warning(self, "Warning", "No files available")
            return
        
        percentage = float(self.percentage_edit.text()) if validate_percentage(self.percentage_edit.text()) else 10.0
        dialog = OutOfRangeFilesDialog(self, file_names, self.crm_db_path, percentage, self.ver_db_path)
        dialog.exec_()

    def edit_record(self):
        selected = self.table_widget.currentRow()
        if selected < 0:
            QMessageBox.warning(self, "Warning", "Please select a record to edit")
            return
        
        all_df = pd.concat([self.filtered_crm_df_cache, self.filtered_blank_df_cache])
        if all_df.empty or selected >= len(all_df):
            QMessageBox.warning(self, "Warning", "No valid record selected")
            return
        
        record = all_df.iloc[selected]
        dialog = EditRecordDialog(self, record, self.crm_db_path)
        if dialog.exec_() == QDialog.Accepted:
            updated_record = dialog.get_updated_record()
            try:
                conn = sqlite3.connect(self.crm_db_path)
                cursor = conn.cursor()
                cursor.execute(
                    """
                    UPDATE crm_data
                    SET crm_id = ?, solution_label = ?, element = ?, value = ?, date = ?
                    WHERE id = ?
                    """,
                    (
                        updated_record['crm_id'],
                        updated_record['solution_label'],
                        updated_record['element'],
                        updated_record['value'],
                        updated_record['date'],
                        record['id']
                    )
                )
                conn.commit()
                conn.close()
                logger.info(f"Updated record ID {record['id']} with new values: {updated_record}")
                self.load_data_thread()
                self.status_label.setText("Record updated successfully")
            except Exception as e:
                logger.error(f"Error updating record: {str(e)}")
                QMessageBox.critical(self, "Error", f"Failed to update record: {str(e)}")

    def on_mouse_clicked(self, event):
        if event.button() == Qt.LeftButton:
            pos = self.plot_widget.getViewBox().mapSceneToView(event.scenePos())
            x, y = pos.x(), pos.y()
            logger.debug(f"Click at view coordinates: x={x:.2f}, y={y:.2f}")
            closest_dist = float('inf')
            closest_info = None

            for plot_item, crm_df, indices, date_labels in self.plot_data_items:
                for i, (idx, value, date) in enumerate(zip(indices, crm_df['value'], date_labels)):
                    dist = ((idx - x) ** 2 + (value - y) ** 2) ** 0.5
                    logger.debug(f"Point {i}: index={idx}, value={value:.2f}, dist={dist:.2f}")
                    if dist < 10:
                        closest_dist = dist
                        element = crm_df.iloc[i]['element']
                        file_name = crm_df.iloc[i]['file_name']
                        folder_name = crm_df.iloc[i]['folder_name']
                        solution_label = crm_df.iloc[i]['solution_label']
                        blank_value = crm_df.iloc[i].get('blank_value')
                        original_value = crm_df.iloc[i].get('original_value', value)

                        blank_info = ""
                        if not self.filtered_blank_df_cache.empty:
                            relevant_blanks = self.filtered_blank_df_cache[
                                (self.filtered_blank_df_cache['file_name'] == file_name) &
                                (self.filtered_blank_df_cache['folder_name'] == folder_name) &
                                (self.filtered_blank_df_cache['element'] == element)
                            ]
                            if not relevant_blanks.empty:
                                blank_info = "\nBLANK Data:\n"
                                for _, blank_row in relevant_blanks.iterrows():
                                    blank_info += f"  - Solution Label: {blank_row['solution_label']}, Value: {blank_row['value']:.2f}\n"

                        closest_info = (
                            f"Element: {element}\n"
                            f"File: {file_name}\n"
                            f"Date: {date}\n"
                            f"Solution Label: {solution_label}\n"
                            f"Value: {value:.2f}\n"
                            f"Original Value: {original_value:.2f}\n" if blank_value is not None else f"Value: {value:.2f}\n"
                            f"Blank Value Applied: {blank_value:.2f}\n" if blank_value is not None else ""
                            f"{blank_info}"
                        )

            if closest_info:
                QMessageBox.information(self, "Point Info", closest_info)
                logger.debug(f"Clicked point info: {closest_info}")
            else:
                logger.debug("No point found near click position")

    def on_mouse_moved(self, pos):
            try:
                pos = self.plot_widget.getViewBox().mapSceneToView(pos)
                x, y = pos.x(), pos.y()
                closest_dist = float('inf')
                closest_info = None
                closest_point = None

                # Get current view range for normalization
                view_box = self.plot_widget.getViewBox()
                x_min, x_max = view_box.viewRange()[0]
                y_min, y_max = view_box.viewRange()[1]
                x_range = x_max - x_min if x_max != x_min else 1
                y_range = y_max - y_min if y_max != y_min else 1

                for plot_item, crm_df, indices, date_labels in self.plot_data_items:
                    plot_data = plot_item.getData()
                    if plot_data is None:
                        continue
                    plot_x, plot_y = plot_data

                    # Normalized dist
                    dx = (plot_x - x) / x_range
                    dy = (plot_y - y) / y_range
                    distances = np.sqrt(dx**2 + dy**2)
                    min_dist_idx = np.argmin(distances)
                    min_dist = distances[min_dist_idx]

                    if min_dist < closest_dist:
                        closest_dist = min_dist
                        i = min_dist_idx
                        value = crm_df.iloc[i]['value']
                        date = date_labels[i]
                        file_name = crm_df.iloc[i]['file_name']
                        folder_name = crm_df.iloc[i]['folder_name']
                        crm_id = crm_df.iloc[i]['norm_crm_id']
                        element = crm_df.iloc[i]['element']
                        solution_label = crm_df.iloc[i]['solution_label']
                        blank_value = crm_df.iloc[i].get('blank_value')
                        original_value = crm_df.iloc[i].get('original_value', value)

                        blank_info = ""
                        if not self.filtered_blank_df_cache.empty:
                            relevant_blanks = self.filtered_blank_df_cache[
                                (self.filtered_blank_df_cache['file_name'] == file_name) &
                                (self.filtered_blank_df_cache['folder_name'] == folder_name) &
                                (self.filtered_blank_df_cache['element'] == element)
                            ]
                            if not relevant_blanks.empty:
                                blank_info = "\nBLANK Data:\n"
                                for _, blank_row in relevant_blanks.iterrows():
                                    blank_info += f"  - {blank_row['solution_label']}: {blank_row['value']:.6f}\n"

                        closest_info = (
                            f"CRM ID: {crm_id}\n"
                            f"Element: {element}\n"
                            f"Date: {date}\n"
                            f"Value: {value:.6f}\n"
                            f"Original Value: {original_value:.6f}\n" if blank_value is not None else f"Value: {value:.6f}\n"
                            f"Blank Value Applied: {blank_value:.6f}\n" if blank_value is not None else ""
                            f"Solution Label: {solution_label}\n"
                            f"File: {file_name}\n"
                            f"{blank_info}"
                        )
                        closest_point = (plot_x[min_dist_idx], plot_y[min_dist_idx])

                if closest_info and closest_dist < 0.05:  # Adjusted normalized threshold
                    self.tooltip_label.setText(closest_info)
                    self.tooltip_label.adjustSize()
                    tooltip_pos = self.plot_widget.getViewBox().mapFromView(pos)
                    self.tooltip_label.move(int(tooltip_pos.x() + 15), int(tooltip_pos.y() - self.tooltip_label.height() / 2))
                    self.tooltip_label.setVisible(True)
                else:
                    self.tooltip_label.setVisible(False)
            except Exception as e:
                logger.error(f"Error in on_mouse_moved: {str(e)}")
                self.tooltip_label.setVisible(False)

    def save_plot(self):
        try:
            import pyqtgraph.exporters
            temp_file = 'temp_crm_plot.png'
            exporter = pyqtgraph.exporters.ImageExporter(self.plot_widget.getPlotItem())
            exporter.parameters()['width'] = 1200
            exporter.export(temp_file)
            im = Image.open(temp_file)
            if self.logo_path.exists():
                logo = Image.open(self.logo_path)
                logo = logo.resize((100, 100))
                box = (im.width - 110, 10)
                if logo.mode == 'RGBA':
                    im.paste(logo, box, logo)
                else:
                    im.paste(logo, box)
                im.save('crm_plot.png')
                os.remove(temp_file)
                self.status_label.setText("Plot saved as crm_plot.png")
                logger.info("Plot saved as crm_plot.png")
        except Exception as e:
            logger.error(f"Error saving plot: {str(e)}")
            self.status_label.setText("Failed to save plot")
            QMessageBox.critical(self, "Error", f"Failed to save plot: {str(e)}")

    def reset_filters(self):
        if self.updating_filters:
            return

        # ریست فیلترها
        if self.device_combo.count() > 0:
            self.device_combo.setCurrentIndex(0)
        if self.crm_combo.count() > 0:
            self.crm_combo.setCurrentIndex(0)
        self.from_date_edit.clear()
        self.to_date_edit.clear()
        self.percentage_edit.setText("10")
        self.best_wl_check.setChecked(True)
        self.apply_blank_check.setChecked(False)

        # ذخیره بعد از ریست
        self.save_settings()
        logger.debug("Filters reset and saved")
        self.update_filters()

    def apply_styles(self):
        self.setStyleSheet("""
            QMainWindow {
                background-color: #F5F6FA;
                font-family: 'Segoe UI', Arial, sans-serif;
            }
            QTableWidget {
                background-color: #FFFFFF;
                border: 1px solid #E0E0E0;
                border-radius: 8px;
                gridline-color: #E0E0E0;
            }
            QTableWidget::item {
                padding: 8px;
            }
            QHeaderView::section {
                background-color: #0078D4;
                color: #FFFFFF;
                border: 1px solid #E0E0E0;
                padding: 8px;
                font-weight: bold;
                font-size: 14px;
            }
            QProgressBar {
                border: 1px solid #E0E0E0;
                border-radius: 4px;
                text-align: center;
                background-color: #FFFFFF;
                color: #000000;
            }
            QProgressBar::chunk {
                background-color: #0078D4;
                border-radius: 4px;
            }
            QLabel {
                color: #000000;
                font-size: 14px;
                font-family: 'Segoe UI';
            }
        """)
        self.plot_widget.setBackground('#FFFFFF')
        self.plot_widget.getAxis('bottom').setPen('#000000')
        self.plot_widget.getAxis('left').setPen('#000000')
        self.plot_widget.getAxis('bottom').setTextPen('#000000')
        self.plot_widget.getAxis('left').setTextPen('#000000')

if __name__ == "__main__":
    app = QApplication(sys.argv)
    # sys.stdout.reconfigure(encoding='utf-8')
    window = CRMDataVisualizer()
    window.show()
    sys.exit(app.exec_())